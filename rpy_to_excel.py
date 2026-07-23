#!/usr/bin/env python3
"""
将 .rpy 文件反向转为 excel_to_rpy.py 对应的 Excel 格式。
用法: python rpy_to_excel.py script.rpy [-o output.xlsx]
"""

import re
import sys
import os
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("请先安装 openpyxl：pip install openpyxl")
    sys.exit(1)

# 复用主脚本的工具函数
try:
    from excel_to_rpy import scan_existing_scripts, _add_dropdown_validation
except ImportError:
    scan_existing_scripts = lambda d=None: (set(), set(), set(), [])
    _add_dropdown_validation = lambda *a, **k: None

# ── 英→中 反向映射 ──────────────────────────────────────

TRANSITION_REVERSE = {
    "dissolve": "溶解", "fade": "褪色", "pixellate": "像素化",
    "hpunch": "横向振动", "vpunch": "纵向振动", "blinds": "百叶窗",
    "squares": "网格覆盖", "wipeleft": "擦除", "slideleft": "滑入",
    "slideawayleft": "滑出", "pushright": "推出",
}

POSITION_REVERSE = {
    "left": "左边", "right": "右边", "center": "中间",
    "truecenter": "真中", "offscreenleft": "左外", "offscreenright": "右外",
}


def _reverse_translate(s: str) -> str:
    """将英文转场/位置翻译为中文"""
    if not s:
        return ""
    parts = s.split()
    result = []
    for p in parts:
        if p in POSITION_REVERSE:
            result.append(POSITION_REVERSE[p])
        elif p in TRANSITION_REVERSE:
            result.append(TRANSITION_REVERSE[p])
        else:
            result.append(p)
    return " ".join(result)


HEADERS = [
    ("场景标签", 16),
    ("指令类型", 20),
    ("图片/背景", 24),
    ("角色名", 14),
    ("变量名", 14),
    ("逻辑连接", 10),   # 新增
    ("对话文本", 40),
    ("选项文本", 20),
    ("跳转目标", 16),
    ("音频路径", 24),
    ("位置/特效/属性", 22),
    ("备注", 20),
]

LOGIC_CONNECTORS = ["and", "or"]

COMMAND_TYPES = [
    "label（场景标签）",
    "scene（背景图）",
    "show（显示角色）",
    "hide（隐藏角色）",
    "dialogue（角色对话）",
    "narrator（旁白）",
    "menu（选择菜单）",
    "menu_option（菜单选项）",
    "jump（跳转）",
    "call（调用子场景）",
    "return（返回）",
    "play_music（播放BGM）",
    "stop_music（停止BGM）",
    "queue_music（排队播BGM）",
    "play_sound（播放音效）",
    "stop_sound（停止音效）",
    "voice（配音）",
    "pause（暂停等待）",
    "player_input（玩家输入）",
    "window（对话框开关）",
    "define_character（定义角色）",
    "define_variable（定义变量）",
    "define_image（定义图片）",
    "variable_set（变量赋值）",
    "variable_change（变量增减）",
    "variable_toggle（变量开关）",
    "variable_eq（变量=）",
    "variable_ne（变量≠）",
    "variable_gt（变量>）",
    "variable_ge（变量≥）",
    "variable_lt（变量<）",
    "variable_le（变量≤）",
    "$（设置变量）",
    "if（条件判断）",
    "elif（否则如果）",
    "else（否则）",
    "default（默认变量）",
    "image（定义图片）",
]


# ── 结构化变量指令反向匹配 ──────────────────────────────

# 结构化条件运算符 → 指令类型
STRUCTURED_COND_MAP = {
    "==": "variable_eq",
    "!=": "variable_ne",
    ">=": "variable_ge",
    "<=": "variable_le",
    ">": "variable_gt",
    "<": "variable_lt",
}

# 结构化条件指令 → 中文下拉标签
STRUCTURED_COND_LABEL = {
    "variable_eq": "variable_eq（变量=）",
    "variable_ne": "variable_ne（变量≠）",
    "variable_gt": "variable_gt（变量>）",
    "variable_ge": "variable_ge（变量≥）",
    "variable_lt": "variable_lt（变量<）",
    "variable_le": "variable_le（变量≤）",
}


def _try_match_structured_condition(content: str) -> dict | None:
    """
    尝试匹配结构化条件：if <var> <op> <val>:
    返回匹配结果 dict 或 None
    """
    m = re.match(r'(if|elif)\s+(\w+)\s*(==|!=|>=|<=|>|<)\s*(\S+?)\s*:\s*$', content)
    if not m:
        return None
    keyword = m.group(1)
    var_name = m.group(2)
    op = m.group(3)
    val = m.group(4)
    if not re.match(r'^-?\d+(\.\d+)?$', val):
        return None
    cmd = STRUCTURED_COND_MAP.get(op)
    if not cmd:
        return None
    label = STRUCTURED_COND_LABEL.get(cmd, f"{cmd}（变量{op}）")
    return {
        "_type": "if" if keyword == "if" else "elif",
        "指令类型": label,
        "变量名": var_name,
        "对话文本": val,
    }


def _try_match_structured_dollar(content: str) -> dict | None:
    """
    尝试匹配结构化 $ 语句：
    - $ var += val  → variable_change
    - $ var -= val  → variable_change
    - $ var = True/False → variable_toggle
    - $ var = val (simple) → variable_set
    返回匹配结果 dict 或 None（复杂表达式回退到 $）
    """
    # variable_change: += 或 -=
    m = re.match(r'\$\s+(\w+)\s*(\+|-)=\s*(.+)$', content)
    if m:
        var_name = m.group(1)
        op = m.group(2)
        val = m.group(3).strip()
        return {
            "_type": "$",
            "指令类型": "variable_change（变量增减）",
            "变量名": var_name,
            "对话文本": f"{op}{val}",
        }

    # variable_toggle: = True 或 = False
    m = re.match(r'\$\s+(\w+)\s*=\s*(True|False)\s*$', content)
    if m:
        var_name = m.group(1)
        val = "true" if m.group(2) == "True" else "false"
        return {
            "_type": "$",
            "指令类型": "variable_toggle（变量开关）",
            "变量名": var_name,
            "对话文本": val,
        }

    # variable_set: = value (简单非函数赋值)
    m = re.match(r'\$\s+(\w+)\s*=\s*(.+)$', content)
    if m:
        var_name = m.group(1)
        val = m.group(2).strip()
        # 只接受简单字面量：数字、标识符、True/False/None、引号字符串
        is_simple = (
            re.match(r'^-?\d+(\.\d+)?$', val)
            or val in ("True", "False", "None")
            or re.match(r'^[A-Za-z_]\w*$', val)
            or (val.startswith('"') and val.endswith('"'))
            or (val.startswith("'") and val.endswith("'"))
        )
        if not is_simple:
            return None
        return {
            "_type": "$",
            "指令类型": "variable_set（变量赋值）",
            "变量名": var_name,
            "对话文本": val,
        }

    return None


def _try_parse_compound_condition(content: str) -> list | None:
    """
    尝试匹配复合结构化条件：if score >= 10 and score < 20:
    返回多个 result dict 的列表（每个子条件一行），或 None。
    """
    m = re.match(r'(if|elif)\s+(.+?)\s*:\s*$', content)
    if not m:
        return None
    keyword = m.group(1)
    body = m.group(2).strip()

    tokens = re.split(r'\s+(and|or)\s+', body)
    if len(tokens) <= 1:
        return None

    results = []
    for i in range(0, len(tokens), 2):
        cond_str = tokens[i].strip()
        connector = tokens[i + 1] if i + 1 < len(tokens) else ""

        m2 = re.match(r'(\w+)\s*(==|!=|>=|<=|>|<)\s*(\S+)', cond_str)
        if not m2:
            return None

        var_name = m2.group(1)
        op = m2.group(2)
        val = m2.group(3)
        cmd = STRUCTURED_COND_MAP.get(op)
        if not cmd:
            return None

        label = STRUCTURED_COND_LABEL.get(cmd, f"{cmd}（变量{op}）")
        result_type = keyword if i == 0 else f"{keyword}_continue"
        results.append({
            "_type": result_type,
            "指令类型": label,
            "变量名": var_name,
            "逻辑连接": connector,
            "对话文本": val,
        })
    return results


HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def parse_rpy(filepath: str) -> list:
    """解析 .rpy 文件，返回行列表，每行是一个 dict"""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        raw_lines = f.readlines()

    result = []
    lines = []
    for line in raw_lines:
        stripped = line.rstrip("\n\r")
        if stripped.strip() == "":
            lines.append(("", 0, ""))
        else:
            indent = len(stripped) - len(stripped.lstrip())
            content = stripped.strip()
            lines.append((stripped, indent, content))

    in_menu = False
    in_python = False
    python_indent = 0
    in_if_body = False
    if_base_indent = -1

    for i, (raw, indent, content) in enumerate(lines):
        if content == "":
            if in_if_body:
                result.append({"_type": "blank"})
            in_menu = False
            in_if_body = False
            continue

        if content.startswith("#"):
            continue

        if in_if_body and indent <= if_base_indent:
            result.append({"_type": "blank"})
            in_if_body = False

        if content.startswith("#"):
            continue

        # python: 块处理
        if in_python:
            if indent <= python_indent and not content.startswith(" "):
                in_python = False
                result.append({"_type": "blank"})
            else:
                # 识别 python 块内的 if/elif/else
                struct = _try_match_structured_condition(content)
                compound = _try_parse_compound_condition(content)
                if struct:
                    result.append(struct)
                elif compound:
                    result.extend(compound)
                elif content == "else:":
                    result.append({"_type": "else", "指令类型": "else（否则）"})
                elif re.match(r"elif\s+", content) and content.rstrip().endswith(":"):
                    result.append({"_type": "elif", "指令类型": "elif（否则如果）",
                                   "对话文本": content[5:].rstrip(":").strip()})
                elif re.match(r"if\s+", content) and content.rstrip().endswith(":"):
                    result.append({"_type": "if", "指令类型": "if（条件判断）",
                                   "对话文本": content[3:].rstrip(":").strip()})
                else:
                    # 尝试结构化 $
                    struct_dollar = _try_match_structured_dollar(content)
                    if struct_dollar:
                        result.append(struct_dollar)
                    else:
                        result.append({"_type": "$", "指令类型": "$（设置变量）",
                                       "对话文本": content})
                continue

        # Top-level define / default / image
        if indent == 0:
            if content.startswith("define "):
                m = re.match(r"define\s+(\w+)\s*=\s*(.+)", content)
                if m:
                    result.append({
                        "_type": "define_character",
                        "指令类型": "define_character（定义角色）",
                        "角色名": m.group(1),
                        "对话文本": m.group(2),
                    })
                continue

            if content.startswith("default "):
                m = re.match(r"default\s+(\w+)\s*=\s*(.+)", content)
                if m:
                    val = m.group(2).strip()
                    result.append({
                        "_type": "define_variable",
                        "指令类型": "define_variable（定义变量）",
                        "变量名": m.group(1),
                        "对话文本": val,
                    })
                continue

            if content.startswith("image "):
                m = re.match(r"image\s+(.+?)\s*=\s*(.+)", content)
                if m:
                    img_name = m.group(1).strip()
                    img_val = m.group(2).strip()
                    if (img_val.startswith('"') and img_val.endswith('"')) or \
                       (img_val.startswith("'") and img_val.endswith("'")):
                        img_val = img_val[1:-1]
                    result.append({
                        "_type": "define_image",
                        "指令类型": "define_image（定义图片）",
                        "变量名": img_name,
                        "图片/背景": img_val,
                    })
                continue

            if content.startswith("label "):
                m = re.match(r"label\s+(\w+)\s*:", content)
                if m:
                    result.append({
                        "_type": "label",
                        "指令类型": "label（场景标签）",
                        "场景标签": m.group(1),
                    })
                in_menu = False
                continue

        # Indented content
        if indent >= 4:
            # menu:
            if re.match(r"menu\s*:", content):
                result.append({
                    "_type": "menu",
                    "指令类型": "menu（选择菜单）",
                })
                in_menu = True
                continue

            # menu option
            if in_menu and re.match(r'"([^"]*)"\s*:', content):
                m = re.match(r'"([^"]*)"\s*:', content)
                result.append({
                    "_type": "menu_option",
                    "指令类型": "menu_option（菜单选项）",
                    "选项文本": m.group(1),
                    "跳转目标": "",
                })
                continue

            if in_menu and result and result[-1]["_type"] == "menu_option":
                if content.startswith("jump "):
                    result[-1]["跳转目标"] = "jump " + content[5:].strip()
                    continue
                elif content == "return":
                    result[-1]["跳转目标"] = "return"
                    continue
                elif content.startswith("call "):
                    result[-1]["跳转目标"] = "call " + content[5:].strip()
                    continue
                elif content.startswith("$ "):
                    result[-1]["跳转目标"] = "$ " + content[2:].strip()
                    continue
                continue

            # 结构化条件（缩进中的 if/elif）
            struct_cond = _try_match_structured_condition(content)
            if struct_cond:
                result.append(struct_cond)
                in_if_body = True
                if_base_indent = indent
                continue

            # 复合结构化条件（if var1 op val1 and var2 op val2:）
            compound_conds = _try_parse_compound_condition(content)
            if compound_conds:
                result.extend(compound_conds)
                in_if_body = True
                if_base_indent = indent
                continue

            # 结构化 $ 语句
            struct_dollar = _try_match_structured_dollar(content)
            if struct_dollar:
                result.append(struct_dollar)
                continue

            # if / elif / else (传统)
            if re.match(r"if\s+", content) and content.rstrip().endswith(":"):
                cond = content[3:].rstrip(":").strip()
                result.append({
                    "_type": "if",
                    "指令类型": "if（条件判断）",
                    "对话文本": cond,
                })
                in_if_body = True
                if_base_indent = indent
                continue

            if re.match(r"elif\s+", content) and content.rstrip().endswith(":"):
                cond = content[5:].rstrip(":").strip()
                result.append({
                    "_type": "elif",
                    "指令类型": "elif（否则如果）",
                    "对话文本": cond,
                })
                in_if_body = True
                if_base_indent = indent
                continue

            if content == "else:":
                result.append({
                    "_type": "else",
                    "指令类型": "else（否则）",
                })
                in_if_body = True
                if_base_indent = indent
                continue

            # python: 块
            if content == "python:":
                in_python = True
                python_indent = indent
                continue

            # scene
            if content.startswith("scene "):
                rest = content[6:].strip()
                img = ""
                effect = ""
                m = re.match(r"(.+?)\s+with\s+(.+)", rest)
                if m:
                    img = m.group(1).strip()
                    effect = m.group(2).strip()
                else:
                    img = rest
                result.append({
                    "_type": "scene",
                    "指令类型": "scene（背景图）",
                    "图片/背景": img,
                    "位置/特效/属性": _reverse_translate(effect),
                })
                continue

            # show
            if content.startswith("show "):
                rest = content[5:].strip()
                img = ""
                effect = ""

                parts = rest.split()
                i = 0
                img_parts = []
                while i < len(parts) and parts[i] not in ("as", "at", "with"):
                    img_parts.append(parts[i])
                    i += 1
                img = " ".join(img_parts)

                remaining = parts[i:]
                effect_parts = []
                skip_next = False
                for p in remaining:
                    if p in ("as", "at", "with"):
                        continue
                    effect_parts.append(p)
                effect = " ".join(effect_parts)

                result.append({
                    "_type": "show",
                    "指令类型": "show（显示角色）",
                    "图片/背景": img,
                    "位置/特效/属性": _reverse_translate(effect),
                })
                continue

            # hide
            if content.startswith("hide "):
                rest = content[5:].strip()
                effect = ""
                m = re.match(r"(\S+)\s+(.+)", rest)
                if m:
                    rest = m.group(1)
                    effect = re.sub(r"\bwith\s+", "", m.group(2)).strip()
                result.append({
                    "_type": "hide",
                    "指令类型": "hide（隐藏角色）",
                    "角色名": rest,
                    "位置/特效/属性": _reverse_translate(effect),
                })
                continue

            # stop music
            if content.startswith("stop music"):
                rest = content[10:].strip()
                effect = rest if rest else ""
                result.append({
                    "_type": "stop_music",
                    "指令类型": "stop_music（停止BGM）",
                    "位置/特效/属性": _reverse_translate(effect),
                })
                continue

            # stop sound
            if content == "stop sound":
                result.append({
                    "_type": "stop_sound",
                    "指令类型": "stop_sound（停止音效）",
                })
                continue

            # play music
            if content.startswith("play music "):
                m = re.match(r'play music\s+"([^"]*)"', content)
                audio = m.group(1) if m else ""
                result.append({
                    "_type": "play_music",
                    "指令类型": "play_music（播放BGM）",
                    "音频路径": audio,
                })
                continue

            # play sound
            if content.startswith("play sound "):
                m = re.match(r'play sound\s+"([^"]*)"', content)
                audio = m.group(1) if m else ""
                result.append({
                    "_type": "play_sound",
                    "指令类型": "play_sound（播放音效）",
                    "音频路径": audio,
                })
                continue

            # queue music
            if content.startswith("queue music "):
                m = re.match(r'queue music\s+"([^"]*)"', content)
                audio = m.group(1) if m else ""
                result.append({
                    "_type": "queue_music",
                    "指令类型": "queue_music（排队播BGM）",
                    "音频路径": audio,
                })
                continue

            # voice
            if content.startswith("voice "):
                m = re.match(r'voice\s+"([^"]*)"', content)
                audio = m.group(1) if m else ""
                result.append({
                    "_type": "voice",
                    "指令类型": "voice（配音）",
                    "音频路径": audio,
                })
                continue

            # window
            if content.startswith("window "):
                action = content[7:].strip()
                result.append({
                    "_type": "window",
                    "指令类型": "window（对话框开关）",
                    "对话文本": action,
                })
                continue

            # pause
            if content.startswith("pause "):
                dur = content[6:].strip()
                result.append({
                    "_type": "pause",
                    "指令类型": "pause（暂停等待）",
                    "对话文本": dur,
                })
                continue

            # return
            if content == "return":
                result.append({
                    "_type": "return",
                    "指令类型": "return（返回）",
                })
                continue

            # jump
            if content.startswith("jump "):
                target = content[5:].strip()
                result.append({
                    "_type": "jump",
                    "指令类型": "jump（跳转）",
                    "跳转目标": target,
                })
                continue

            # call
            if content.startswith("call "):
                target = content[5:].strip()
                result.append({
                    "_type": "call",
                    "指令类型": "call（调用子场景）",
                    "跳转目标": target,
                })
                continue

            # $ (python one-liner) — fallback for remaining
            if content.startswith("$ "):
                code = content[2:].strip()
                result.append({
                    "_type": "$",
                    "指令类型": "$（设置变量）",
                    "对话文本": code,
                })
                continue

            # centered
            if content.startswith("centered "):
                m = re.match(r'centered\s+"([^"]*)"', content)
                text = m.group(1) if m else content[9:].strip().strip('"')
                result.append({
                    "_type": "narrator",
                    "指令类型": "narrator（旁白）",
                    "对话文本": text,
                    "备注": "centered",
                })
                continue

            # narrator "text"
            if content.startswith("narrator "):
                m = re.match(r'narrator\s+"([^"]*)"', content)
                text = m.group(1) if m else content[9:].strip().strip('"')
                result.append({
                    "_type": "narrator",
                    "指令类型": "narrator（旁白）",
                    "对话文本": text,
                    "备注": "explicit",
                })
                continue

            # character "text" (dialogue)
            m_dialogue = re.match(r'(\w+)\s+"([^"]*)"', content)
            if m_dialogue:
                char = m_dialogue.group(1)
                text = m_dialogue.group(2)
                if char in ("narrator", "centered", ""):
                    result.append({
                        "_type": "narrator",
                        "指令类型": "narrator（旁白）",
                        "对话文本": text,
                    })
                else:
                    result.append({
                        "_type": "dialogue",
                        "指令类型": "dialogue（角色对话）",
                        "角色名": char,
                        "对话文本": text,
                    })
                continue

            # bare string "text" (narrator)
            m_bare = re.match(r'"([^"]*)"', content)
            if m_bare:
                result.append({
                    "_type": "narrator",
                    "指令类型": "narrator（旁白）",
                    "对话文本": m_bare.group(1),
                })
                rest = content[m_bare.end():].strip()
                m2 = re.match(r'"([^"]*)"', rest)
                if m2:
                    result.append({
                        "_type": "narrator",
                        "指令类型": "narrator（旁白）",
                        "对话文本": m2.group(1),
                    })
                continue

    return result


def _write_sheet(wb, sheet_name: str, items: list, dropdowns=None):
    """将 items 写入一个 Sheet（12 列）"""
    if sheet_name in [s.title for s in wb.worksheets]:
        ws = wb[sheet_name]
    elif wb.worksheets and wb.worksheets[0].title == "Sheet":
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    _setup_sheet_header(ws, dropdowns)

    row_num = 1
    for item in items:
        row_num += 1
        if item.get("_type") == "blank":
            for col_idx in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col_idx, value="")
                cell.border = THIN_BORDER
            ws.row_dimensions[row_num].height = 8
            continue

        values = [
            item.get("场景标签", ""),
            item.get("指令类型", ""),
            item.get("图片/背景", ""),
            item.get("角色名", ""),
            item.get("变量名", ""),
            item.get("逻辑连接", ""),       # 新增列
            item.get("对话文本", ""),
            item.get("选项文本", ""),
            item.get("跳转目标", ""),
            item.get("音频路径", ""),
            item.get("位置/特效/属性", ""),
            item.get("备注", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            if row_num % 2 == 0:
                cell.fill = ALT_FILL
        ws.row_dimensions[row_num].height = 20

    ws.freeze_panes = "A2"


def _setup_sheet_header(ws, dropdowns=None):
    """设置 Sheet 表头（仅当表头不存在时），并添加下拉验证"""
    if ws.cell(1, 1).value:
        return
    for col_idx, (header, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24

    # 指令类型下拉（列表较长时自动使用隐藏辅助 Sheet）
    _add_dropdown_validation(ws, "B", COMMAND_TYPES)

    # 逻辑连接下拉（F 列）
    dv_connect = DataValidation(type="list", formula1='"and,or"', allow_blank=True)
    dv_connect.add("F2:F10000")
    ws.add_data_validation(dv_connect)

    # 变量开关下拉（G 列）
    dv_toggle = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
    dv_toggle.add("G2:G10000")
    ws.add_data_validation(dv_toggle)

    if dropdowns:
        if dropdowns.get("characters"):
            _add_dropdown_validation(ws, "D", dropdowns["characters"])
        if dropdowns.get("variables"):
            _add_dropdown_validation(ws, "E", dropdowns["variables"])
        if dropdowns.get("images"):
            _add_dropdown_validation(ws, "C", dropdowns["images"])


def write_excel(rows: list, output_path: str, split_sheets: bool = True):
    """将解析结果写入 Excel，默认按 label 拆分为多 Sheet"""
    wb = Workbook()

    # 从解析结果中收集角色、变量、图片引用
    dd_characters = set()
    dd_variables = set()
    dd_images = set()
    for item in rows:
        if item.get("角色名"):
            dd_characters.add(item["角色名"])
        if item.get("变量名"):
            dd_variables.add(item["变量名"])
        if item.get("图片/背景"):
            v = item["图片/背景"]
            if not v.startswith('"') and not v.startswith("'"):
                dd_images.add(v)

    # 合并 game/ 目录扫描结果
    sc_chars, sc_vars, sc_imgs, sc_files = scan_existing_scripts()
    dd_characters.update(sc_chars)
    dd_variables.update(sc_vars)
    dd_images.update(sc_imgs)
    dd_images.update(sc_files)

    dropdowns = {
        "characters": dd_characters,
        "variables": dd_variables,
        "images": dd_images,
    }

    if not split_sheets:
        _write_sheet(wb, "script", rows, dropdowns)
    else:
        top_items = []
        label_groups = []
        current_label = None
        current_items = []

        for item in rows:
            if item.get("_type") == "label":
                if current_label is not None:
                    label_groups.append((current_label, current_items))
                elif current_items:
                    top_items = current_items
                current_label = item.get("场景标签", "unknown")
                current_items = [item]
            else:
                if current_label is None:
                    top_items.append(item)
                else:
                    current_items.append(item)

        if current_label is not None:
            label_groups.append((current_label, current_items))
        elif current_items:
            top_items = current_items

        sheet_count = 0
        if top_items or label_groups:
            first_name = label_groups[0][0] if label_groups else "script"
            first_items = top_items + (label_groups[0][1] if label_groups else [])
            _write_sheet(wb, first_name, first_items, dropdowns)
            sheet_count += 1

        for label_name, items in label_groups[1:]:
            _write_sheet(wb, label_name, items, dropdowns)
            sheet_count += 1

    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    visible_sheets = sum(1 for ws in wb.worksheets if ws.sheet_state == "visible")
    print(f"   {sum(1 for r in rows if r.get('_type') != 'blank')} rows / {visible_sheets} sheet(s)")


def _strip_quotes(s: str) -> str:
    s = s.strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
        s = s[1:-1]
    return s


def _interactive_convert():
    print("\n" + "-" * 40)
    path = input("请输入 .rpy 文件路径（可直接拖拽文件到此处）：\n> ").strip()
    path = _strip_quotes(path)
    if not path:
        print("[ERROR] 未输入路径，返回主菜单。")
        return
    if not os.path.isfile(path):
        print(f"[ERROR] 文件不存在：{path}")
        return
    if not path.lower().endswith(".rpy"):
        print(f"[ERROR] 不是 .rpy 文件，请拖拽 .rpy 脚本。")
        print(f"       .xlsx 文件请用 excel_to_rpy.py 打开。")
        return
    default_out = str(Path(path).with_suffix(".xlsx"))
    out = input(f"输出路径（直接回车使用默认：{default_out}）：\n> ").strip()
    out = _strip_quotes(out) if out else default_out
    try:
        rows = parse_rpy(path)
    except UnicodeDecodeError:
        print(f"[ERROR] 文件编码错误，请确认是 .rpy 文本文件而非 .xlsx 二进制文件。")
        return
    split = input("按 label 拆分为多 Sheet？(Y/n)：").strip().lower()
    write_excel(rows, out, split_sheets=(split != "n"))


def interactive_mode():
    while True:
        print()
        print("=" * 40)
        print("  Ren'Py .rpy -> Excel 反向转换工具")
        print("=" * 40)
        print("  1. 转换 .rpy 为 Excel")
        print("  2. 退出")
        print("-" * 40)
        choice = input("请选择 (1/2)：").strip()
        if choice == "1":
            _interactive_convert()
            input("\n按 Enter 键继续...")
        elif choice == "2":
            print("再见！")
            break
        else:
            print("[ERROR] 无效选择，请输入 1 或 2。")


def _fix_working_directory():
    """如果当前工作目录是系统目录（如 C:\\Windows\\System32），切换到脚本所在目录。"""
    cwd = os.getcwd()
    windir = os.environ.get("SystemRoot", r"C:\Windows")
    windir = os.path.normpath(windir).lower()
    if os.path.normpath(cwd).lower().startswith(windir + os.sep) or os.path.normpath(cwd).lower() == windir:
        script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        os.chdir(script_dir)
        print(f"[WARN] Switched from system dir ({cwd}) to script dir: {script_dir}\n")


def main():
    _fix_working_directory()

    if len(sys.argv) < 2:
        try:
            interactive_mode()
        except KeyboardInterrupt:
            print("\n\n再见！")
        except Exception as e:
            print(f"\n[ERROR] {e}")
            input("\n按 Enter 键退出...")
        return

    input_path = sys.argv[1]
    output_path = None
    if "-o" in sys.argv:
        idx = sys.argv.index("-o")
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    if not output_path:
        output_path = str(Path(input_path).with_suffix(".xlsx"))

    if not os.path.isfile(input_path):
        print(f"[ERROR] File not found: {input_path}")
        sys.exit(1)

    no_split = "--no-split" in sys.argv
    rows = parse_rpy(input_path)
    write_excel(rows, output_path, split_sheets=not no_split)


if __name__ == "__main__":
    main()
