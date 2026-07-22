#!/usr/bin/env python3
"""
Ren'Py Excel 转 .rpy 脚本 — 新手友好版

用法:
  python excel_to_rpy.py --template         生成 Excel 模板文件
  python excel_to_rpy.py input.xlsx         转换 Excel 为 .rpy
  python excel_to_rpy.py input.xlsx -o out  指定输出路径
  双击打开 → 交互菜单模式
"""

import argparse
import os
import sys
import re
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("请先安装 openpyxl：pip install openpyxl")
    sys.exit(1)

# ── 列索引常量 ──────────────────────────────────────────────
COL_SCENE_LABEL = 0
COL_CMD = 1
COL_IMAGE = 2
COL_CHARACTER = 3
COL_VARIABLE = 4     # 变量名
COL_CONNECT = 5      # 逻辑连接（新增，第6列）
COL_DIALOGUE = 6     # 对话文本
COL_OPTION = 7       # 选项文本
COL_JUMP = 8         # 跳转目标
COL_AUDIO = 9        # 音频路径
COL_EFFECT = 10      # 位置/特效/属性
COL_NOTES = 11       # 备注

# ── 模板配置 ──────────────────────────────────────────────

HEADERS = [
    ("场景标签", 16),
    ("指令类型", 20),
    ("图片/背景", 24),
    ("角色名", 14),
    ("变量名", 14),
    ("逻辑连接", 10),   # 新增：and / or / 空
    ("对话文本", 40),
    ("选项文本", 20),
    ("跳转目标", 16),
    ("音频路径", 24),
    ("位置/特效/属性", 22),
    ("备注", 20),
]

# 逻辑连接下拉选项
LOGIC_CONNECTORS = ["and", "or"]

# 结构化变量条件运算符映射
VARIABLE_OP_MAP = {
    "variable_eq": "==",
    "variable_ne": "!=",
    "variable_gt": ">",
    "variable_ge": ">=",
    "variable_lt": "<",
    "variable_le": "<=",
}

# 变量开关下拉选项
VARIABLE_TOGGLE_VALUES = ["true", "false"]

# 指令类型下拉选项
COMMAND_TYPES = [
    # 原有指令（保持不变）
    "label（场景标签）",
    "scene（背景图）",
    "show（显示角色）",
    "hide（隐藏角色）",
    "dialogue（角色对话）",
    "narrator（旁白）",
    "menu（选择菜单）",
    "menu_option（菜单选项）",
    "jump（跳转）",
    "call（调用子场景）",
    "return（返回）",
    "play_music（播放BGM）",
    "stop_music（停止BGM）",
    "queue_music（排队播BGM）",
    "play_sound（播放音效）",
    "stop_sound（停止音效）",
    "voice（配音）",
    "pause（暂停等待）",
    "player_input（玩家输入）",
    "window（对话框开关）",
    # 定义类
    "define_character（定义角色）",
    "define_variable（定义变量）",
    "define_image（定义图片）",
    # 操作类
    "variable_set（变量赋值）",
    "variable_change（变量增减）",
    "variable_toggle（变量开关）",
    # 条件类（结构化）
    "variable_eq（变量=）",
    "variable_ne（变量≠）",
    "variable_gt（变量>）",
    "variable_ge（变量≥）",
    "variable_lt（变量<）",
    "variable_le（变量≤）",
    # 传统 if/elif/else 和 $ 保留
    "$（设置变量）",
    "if（条件判断）",
    "elif（否则如果）",
    "else（否则）",
    # 向后兼容旧指令
    "default（默认变量）",
    "image（定义图片）",
]

# 兼容旧指令 → 新指令的映射（rpy_to_excel 反向解析用）
OLD_CMD_ALIASES = {
    "default": "define_variable",
    "image": "define_image",
}

# 样式
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


# ── 扫描已有脚本 ──────────────────────────────────────────

def _find_project_root() -> Path:
    """从脚本所在目录向上查找包含 game/ 目录的项目根目录"""
    script_dir = Path(__file__).resolve().parent
    for p in [script_dir] + list(script_dir.parents):
        if (p / "game").is_dir():
            return p
    return Path.cwd()


def scan_existing_scripts(game_dir: str = None):
    """
    扫描 game/ 目录下的 .rpy 文件和图片文件，提取已有定义。

    会自动从脚本目录向上查找项目根目录（含 game/ 的目录）。
    也可通过 game_dir 参数手动指定。

    返回: (characters, variables, image_defs, image_files)
        - characters: set, define 定义的角色名
        - variables: set, default 定义的变量名
        - image_defs: set, image 定义的图片标识名
        - image_files: list, game/ 下找到的图片文件路径（无 image 定义的）
    """
    characters = set()
    variables = set()
    image_defs = {}  # 定义名 → 文件路径

    if game_dir is None:
        root = _find_project_root()
        game_path = root / "game"
    else:
        game_path = Path(game_dir)
        if not game_path.is_dir():
            root = _find_project_root()
            game_path = root / "game"

    if not game_path.is_dir():
        return characters, variables, set(), []

    rpy_files = list(game_path.glob("*.rpy")) + list(game_path.rglob("*.rpy"))

    for rpy_file in rpy_files:
        try:
            content = rpy_file.read_text(encoding="utf-8-sig")
        except Exception:
            continue
        for line in content.splitlines():
            line = line.strip()
            if line.startswith("#"):
                continue
            # define 角色名
            m = re.match(r'define\s+(\w+)\s*=\s*Character', line)
            if m:
                characters.add(m.group(1))
            # default 变量名
            m = re.match(r'default\s+(\w+)\s*=', line)
            if m:
                variables.add(m.group(1))
            # image 定义
            m = re.match(r'image\s+([^=\s]+(?:\s+[^=\s]+)*)\s*=\s*(.+)', line)
            if m:
                img_name = m.group(1).strip()
                img_val = m.group(2).strip()
                if (img_val.startswith('"') and img_val.endswith('"')) or \
                   (img_val.startswith("'") and img_val.endswith("'")):
                    img_val = img_val[1:-1]
                image_defs[img_name] = img_val

    # 扫描图片文件
    image_extensions = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
    image_files = []
    defined_paths = set(image_defs.values())
    defined_paths.update(image_defs.keys())

    gui_dirs = {"gui", "GUI"}
    prefix = str(game_path).replace("\\", "/") + "/"
    for ext in image_extensions:
        for img_file in game_path.glob(f"**/*{ext}"):
            rel = str(img_file).replace("\\", "/")
            parts = Path(rel).parts
            if any(p in gui_dirs for p in parts):
                continue
            # 去掉 game/ 前缀，与 image 定义中的路径格式一致
            if rel.startswith(prefix):
                rel = rel[len(prefix):]
            if rel not in defined_paths:
                image_files.append(rel)

    return characters, variables, set(image_defs.keys()), image_files


def _add_dropdown_validation(ws, col_letter, items, row_range="2:10000"):
    """添加下拉验证，列表过长时写入隐藏辅助 Sheet"""
    if not items:
        return
    sorted_items = sorted(str(i) for i in items)
    # 大约估算：逗号分隔 + 引号
    combined = '"' + ",".join(sorted_items) + '"'
    start_row, end_row = row_range.split(":")
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    if len(combined) <= 255:
        dv = DataValidation(type="list", formula1=combined, allow_blank=True)
        dv.add(cell_range)
        ws.add_data_validation(dv)
    else:
        wb = ws.parent
        helper_name = f"_dd_{col_letter.replace('$', '')}"
        if helper_name in wb.sheetnames:
            del wb[helper_name]
        helper = wb.create_sheet(helper_name)
        helper.sheet_state = "hidden"
        for i, item in enumerate(sorted_items, start=1):
            helper.cell(row=i, column=1, value=item)
        dv = DataValidation(
            type="list",
            formula1=f"={helper_name}!$A$1:$A${len(sorted_items)}",
            allow_blank=True,
        )
        dv.add(cell_range)
        ws.add_data_validation(dv)


def _setup_sheet_header_and_dropdowns(
    ws, scanned_characters, scanned_variables, scanned_images, scanned_files
):
    """为单个 Sheet 写入表头并添加下拉验证"""
    # 表头
    for col_idx, (header, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24

    # 指令类型下拉（列 B → COL_CMD+1=2）
    cmd_formula = '"' + ",".join(COMMAND_TYPES) + '"'
    dv = DataValidation(type="list", formula1=cmd_formula, allow_blank=True)
    dv.error = "请从下拉列表选择指令类型"
    dv.errorTitle = "无效指令"
    dv.prompt = "请选择指令类型（括号内是中文说明）"
    dv.promptTitle = "指令类型"
    dv.add("B2:B10000")
    ws.add_data_validation(dv)

    # 角色名下拉（列 D → COL_CHARACTER+1=4）
    if scanned_characters:
        _add_dropdown_validation(ws, "D", scanned_characters)

    # 变量名下拉（列 E → COL_VARIABLE+1=5）
    if scanned_variables:
        _add_dropdown_validation(ws, "E", scanned_variables)

    # 逻辑连接下拉（列 F → COL_CONNECT+1=6）
    connect_formula = '"and,or"'
    dv_connect = DataValidation(type="list", formula1=connect_formula, allow_blank=True)
    dv_connect.error = "请选择 and 或 or"
    dv_connect.errorTitle = "无效值"
    dv_connect.add("F2:F10000")
    ws.add_data_validation(dv_connect)

    # 图片下拉（列 C → COL_IMAGE+1=3）：image 定义名 + 文件路径
    all_image_refs = set(scanned_images)
    all_image_refs.update(scanned_files)
    if all_image_refs:
        _add_dropdown_validation(ws, "C", all_image_refs)

    # 变量开关值下拉（列 G 当指令类型为 variable_toggle 时使用）
    toggle_formula = '"true,false"'
    dv_toggle = DataValidation(type="list", formula1=toggle_formula, allow_blank=True)
    dv_toggle.error = "请选择 true 或 false"
    dv_toggle.errorTitle = "无效值"
    dv_toggle.add("G2:G10000")
    ws.add_data_validation(dv_toggle)

    ws.freeze_panes = "A2"


# ── 模板生成 ──────────────────────────────────────────────

def _pad_row_for_connect(row):
    """在位置 COL_CONNECT(5) 插入空字符串（逻辑连接列），旧11列→新12列。
    已为12列的跳过不处理。"""
    r = list(row)
    if len(r) >= len(HEADERS):
        return r
    while len(r) < COL_CONNECT:
        r.append("")
    r.insert(COL_CONNECT, "")
    return r


def _get_template_data(scanned_characters, scanned_variables, scanned_images, scanned_files):
    """返回模板数据（所有 Sheet 的行数据），供 generate_template 和 generate_blank_template 共用
    注：每行仍为 11 列（旧格式），_add_sheet 会自动插入逻辑连接列。
    """

    # ── Sheet 1: 入门演示 ──
    sheet_basic = [
        ["start", "label（场景标签）", "", "", "", "", "", "", "", "", "游戏开始"],
        ["", "scene（背景图）", "images/bg_classroom.jpg", "", "", "", "", "", "", "溶解", "场景渐变（溶解=dissolve）"],
        ["", "play_music（播放BGM）", "", "", "", "", "", "audio/bgm_happy.ogg", "", "", "播放背景音乐"],
        ["", "narrator（旁白）", "", "", "", "你醒来发现自己在一间教室里……", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "周围很安静，阳光从窗户洒进来。", "", "", "", "", ""],
        ["", "show（显示角色）", "", "主角", "", "", "", "", "", "左边", "角色出现在左边（角色名可留空沿用上一行）"],
        ["", "dialogue（角色对话）", "", "主角", "", "这里就是新学校吗？", "", "", "", "", ""],
        ["", "dialogue（角色对话）", "", "", "", "得先去教室报到才行。", "", "", "", "", "角色名留空=沿用上一行（主角）"],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "show（显示角色）", "", "同学A", "", "", "", "", "", "右边", "第二个角色入场"],
        ["", "dialogue（角色对话）", "", "同学A", "", "你是新来的吗？", "", "", "", "", ""],
        ["", "dialogue（角色对话）", "", "", "", "我叫小美，欢迎你！", "", "", "", "", ""],
        ["", "hide（隐藏角色）", "", "同学A", "", "", "", "", "", "", "同学A退场"],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "dialogue（角色对话）", "", "主角", "", "看来这里的人都很友善呢。", "", "", "", "", ""],
        ["", "stop_music（停止BGM）", "", "", "", "", "", "", "fadeout 2.0", "", "音乐淡出2秒"],
        ["", "jump（跳转）", "", "", "", "", "", "menu_demo", "", "", "", "跳到选择分支演示"],
    ]

    # ── Sheet 2: 选择分支 ──
    sheet_menu = [
        ["menu_demo", "label（场景标签）", "", "", "", "", "", "", "", "", "选择分支演示"],
        ["", "dialogue（角色对话）", "", "主角", "", "放学后要做什么呢？", "", "", "", "", ""],
        ["", "menu（选择菜单）", "", "", "", "", "", "", "", "", "弹出选项"],
        ["", "menu_option（菜单选项）", "", "", "", "", "去图书馆", "jump library_scene", "", "", ""],
        ["", "menu_option（菜单选项）", "", "", "", "", "去操场", "jump playground_scene", "", "", ""],
        ["", "menu_option（菜单选项）", "", "", "", "", "直接回家", "return", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["library_scene", "label（场景标签）", "", "", "", "", "", "", "", "", "图书馆场景"],
        ["", "scene（背景图）", "images/bg_library.jpg", "", "", "", "", "", "", "褪色", "褪色=fade"],
        ["", "dialogue（角色对话）", "", "主角", "", "今天读了不少书，很有收获。", "", "", "", "", ""],
        ["", "variable_change（变量增减）", "", "", "knowledge", "1", "", "", "", "", "变量 knowledge +1"],
        ["", "variable_ge（变量≥）", "", "", "knowledge", "3", "", "", "", "", "如果 knowledge ≥ 3"],
        ["", "dialogue（角色对话）", "", "主角", "", "我已经读了很多书了！", "", "", "", "", "条件成立时执行"],
        ["", "jump（跳转）", "", "", "", "", "", "menu_demo", "", "", ""],
        ["", "else（否则）", "", "", "", "", "", "", "", "", "条件不成立"],
        ["", "dialogue（角色对话）", "", "主角", "", "再多读一会儿吧~", "", "", "", "", "条件不成立时执行"],
        ["", "jump（跳转）", "", "", "", "", "", "library_scene", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["playground_scene", "label（场景标签）", "", "", "", "", "", "", "", "", "操场场景"],
        ["", "scene（背景图）", "images/bg_playground.jpg", "", "", "", "", "", "", "闪白", "闪白=Fade(0.1,0,0.5)"],
        ["", "dialogue（角色对话）", "", "主角", "", "运动完真舒服！", "", "", "", "", ""],
        ["", "variable_change（变量增减）", "", "", "health", "1", "", "", "", "", ""],
        ["", "jump（跳转）", "", "", "", "", "", "menu_demo", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["multi_check", "label（场景标签）", "", "", "", "", "", "", "", "", "多重条件校验演示"],
        ["", "narrator（旁白）", "", "", "", "多重条件判断：", "", "", "", "", ""],
        ["", "variable_set（变量赋值）", "", "", "score", "80", "", "", "", "", "结构化变量赋值"],
        ["", "variable_set（变量赋值）", "", "", "love", "75", "", "", "", "", ""],
        ["", "if（条件判断）", "", "", "", "score >= 80 and love >= 70", "", "", "", "", "复杂条件仍然用 if"],
        ["", "dialogue（角色对话）", "", "主角", "", "完美！好感度和分数都够了！", "", "", "", "", ""],
        ["", "elif（否则如果）", "", "", "", "score >= 80 or love >= 70", "", "", "", "", "复杂条件用 elif"],
        ["", "dialogue（角色对话）", "", "主角", "", "还差一点点……", "", "", "", "", ""],
        ["", "else（否则）", "", "", "", "", "", "", "", "", "都不达标→坏结局"],
        ["", "dialogue（角色对话）", "", "主角", "", "看来要多努力了……", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "复杂条件写法（传统 if，保留）：", "", "", "", "", ""],
        ["", "if（条件判断）", "", "", "", "has_key and not is_locked", "", "", "", "", ""],
        ["", "if（条件判断）", "", "", "", "money >= 100 or has_discount", "", "", "", "", ""],
        ["", "jump（跳转）", "", "", "", "", "", "menu_demo", "", "", ""],
    ]

    # ── Sheet 3: 音频控制 ──
    sheet_audio = [
        ["audio_demo", "label（场景标签）", "", "", "", "", "", "", "", "", "音频演示"],
        ["", "play_music（播放BGM）", "", "", "", "", "", "audio/bgm_calm.ogg", "", "", "播放BGM"],
        ["", "narrator（旁白）", "", "", "", "背景音乐响起……", "", "", "", "", ""],
        ["", "queue_music（排队播BGM）", "", "", "", "", "", "audio/bgm_tense.ogg", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "气氛渐渐紧张起来。", "", "", "", "", ""],
        ["", "stop_music（停止BGM）", "", "", "", "", "", "", "fadeout 1.0", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "play_sound（播放音效）", "", "", "", "", "", "audio/sfx_door.ogg", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "咚咚咚。", "", "", "", "", ""],
        ["", "stop_sound（停止音效）", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "voice（配音）", "", "", "", "", "", "audio/voice/line01.ogg", "", "", ""],
        ["", "dialogue（角色对话）", "", "主角", "", "你好，我叫小明。", "", "", "", "", ""],
    ]

    # ── Sheet 4: 特效演示 ──
    sheet_effects = [
        ["effect_demo", "label（场景标签）", "", "", "", "", "", "", "", "", "特效/位置中文演示"],
        ["", "scene（背景图）", "images/bg_room.jpg", "", "", "", "", "", "", "溶解", ""],
        ["", "narrator（旁白）", "", "", "", "常用转场：", "", "", "", "", ""],
        ["", "scene（背景图）", "images/bg_room.jpg", "", "", "", "", "", "", "褪色", ""],
        ["", "scene（背景图）", "black", "", "", "", "", "", "", "闪白", ""],
        ["", "scene（背景图）", "images/bg_room.jpg", "", "", "", "", "", "", "像素化", ""],
        ["", "scene（背景图）", "images/bg_room.jpg", "", "", "", "", "", "", "横向振动", ""],
        ["", "scene（背景图）", "images/bg_room.jpg", "", "", "", "", "", "", "纵向振动", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "常用位置：", "", "", "", "", ""],
        ["", "show（显示角色）", "", "主角", "", "", "", "", "", "左边", ""],
        ["", "show（显示角色）", "", "主角", "", "", "", "", "", "右边", ""],
        ["", "show（显示角色）", "", "主角", "", "", "", "", "", "中间", ""],
        ["", "show（显示角色）", "", "主角", "", "", "", "", "", "左外", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "组合使用：", "", "", "", "", ""],
        ["", "show（显示角色）", "", "同学A", "", "", "", "", "", "右边 溶解", "位置+转场 空格分隔"],
        ["", "show（显示角色）", "", "主角", "", "", "", "", "", "左边 褪色", "英文也支持：at left with dissolve"],
    ]

    # ── Sheet 5: 变量与输入 ──
    sheet_vars = [
        ["var_demo", "label（场景标签）", "", "", "", "", "", "", "", "", "结构化变量和输入演示"],
        ["", "define_variable（定义变量）", "", "", "score", "0", "", "", "", "", "结构化变量定义"],
        ["", "define_variable（定义变量）", "", "", "has_key", "False", "", "", "", "", ""],
        ["", "define_variable（定义变量）", "", "", "money", "100", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "让我们来设置一些变量。", "", "", "", "", ""],
        ["", "variable_change（变量增减）", "", "", "score", "10", "", "", "", "", "变量 +10"],
        ["", "variable_set（变量赋值）", "", "", "money", "200", "", "", "", "", "变量赋值 200"],
        ["", "variable_toggle（变量开关）", "", "", "has_key", "true", "", "", "", "", "开关设为 true"],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "player_input（玩家输入）", "", "player_name", "", "请输入你的名字：", "", "", "无名", "", ""],
        ["", "dialogue（角色对话）", "", "主角", "", "你好，[player_name]！", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "", "结构化条件判断：", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "", "【复合条件】score ≥ 10 and score < 50：", "", "", "", "", ""],
        ["", "variable_ge（变量≥）", "", "", "score", "and", "10", "", "", "", "", "第1行：逻辑连接填 and"],
        ["", "variable_lt（变量<）", "", "", "score", "", "50", "", "", "", "", "第2行：逻辑连接为空，结束条件"],
        ["", "dialogue（角色对话）", "", "主角", "", "", "分数在 10~49 之间！", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "", "【复合条件】score == 100 or health <= 0：", "", "", "", "", ""],
        ["", "variable_eq（变量=）", "", "", "score", "or", "100", "", "", "", "", "第1行：or 连接"],
        ["", "variable_le（变量≤）", "", "", "health", "", "0", "", "", "", "", "第2行：结束"],
        ["", "dialogue（角色对话）", "", "主角", "", "", "满分或没血了！", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "", "【单条件】score ≥ 50（逻辑连接留空=独立条件）：", "", "", "", "", ""],
        ["", "variable_ge（变量≥）", "", "", "score", "", "50", "", "", "", "", ""],
        ["", "dialogue（角色对话）", "", "主角", "", "", "及格了！", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "call（调用子场景）", "", "", "", "", "", "audio_demo", "", "", ""],
        ["", "return（返回）", "", "", "", "", "", "", "", "", ""],
    ]

    # ── Sheet 6: 高级技巧 ──
    sheet_advanced = [
        ["advanced", "label（场景标签）", "", "", "", "", "", "", "", "", "高级技巧演示"],
        ["", "scene（背景图）", "black", "", "", "", "", "", "", "", ""],
        ["", "window（对话框开关）", "", "", "", "hide", "", "", "", "", "隐藏对话框"],
        ["", "narrator（旁白）", "", "", "", "第一章", "", "", "", "", "centered"],
        ["", "pause（暂停等待）", "", "", "", "2.0", "", "", "", "", ""],
        ["", "narrator（旁白）", "", "", "", "启程之日", "", "", "", "", "centered"],
        ["", "pause（暂停等待）", "", "", "", "2.0", "", "", "", "", ""],
        ["", "window（对话框开关）", "", "", "", "show", "", "", "", "", "恢复对话框"],
        ["", "jump（跳转）", "", "", "", "", "", "start", "", "", ""],
    ]

    # ── Sheet 7: 角色与图片定义 ──
    sheet_defines = [
        ["", "define_character（定义角色）", "", "主角", "", 'Character("小明", color="#4a90d9")', "", "", "", "", "定义角色名和颜色"],
        ["", "define_character（定义角色）", "", "同学A", "", "", "", "", "", "", "省略参数=默认 Character"],
        ["", "define_variable（定义变量）", "", "", "money", "100", "", "", "", "", "结构化变量定义"],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "define_image（定义图片）", "images/bg_room.jpg", "", "bg_room", "", "", "", "", "", "图片定义：列C填路径，列E填标识名"],
        ["", "define_image（定义图片）", "images/hero_happy.png", "", "主角 happy", "", "", "", "", "", "带属性图片"],
        ["", "define_image（定义图片）", 'Transform("images/bg.jpg", size=(1920,1080))', "", "bg_custom", "", "", "", "", "", "函数调用直接写"],
    ]

    sheets = [
        ("入门演示", sheet_basic),
        ("选择分支", sheet_menu),
        ("音频控制", sheet_audio),
        ("特效演示", sheet_effects),
        ("变量与输入", sheet_vars),
        ("高级技巧", sheet_advanced),
        ("角色与图片定义", sheet_defines),
    ]
    return sheets


def generate_template(output_path: str):
    """生成 Excel 模板文件（多 Sheet 分功能演示）"""
    wb = Workbook()

    scanned_characters, scanned_variables, scanned_images, scanned_files = scan_existing_scripts()

    def _add_sheet(title: str, data_rows: list):
        if wb.worksheets[0].title == "Sheet":
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title)
        _setup_sheet_header_and_dropdowns(
            ws, scanned_characters, scanned_variables, scanned_images, scanned_files
        )
        # 数据
        for row_idx, row_data in enumerate(data_rows, start=2):
            row_data = _pad_row_for_connect(row_data)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGN
                cell.border = THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
            ws.row_dimensions[row_idx].height = 20

    sheets = _get_template_data(
        scanned_characters, scanned_variables, scanned_images, scanned_files
    )
    for sheet_name, data in sheets:
        _add_sheet(sheet_name, data)

    # ── 使用说明 Sheet ──
    ws_help = wb.create_sheet("使用说明")
    help_lines = [
        "═══════════════════════════════════════════",
        "  Ren'Py Excel → .rpy 转换工具 · 使用指南",
        "═══════════════════════════════════════════",
        "",
        "【新增功能（v2）】结构化变量 + 图片管理",
        "  11 列表格（原 10 列 + 新增「变量名」列）：",
        "    A 场景标签  B 指令类型  C 图片/背景  D 角色名",
        "    E 变量名    F 对话文本  G 选项文本  H 跳转目标",
        "    I 音频路径  J 位置/特效  K 备注",
        "",
        "  · 角色名列：只填角色（dialogue / show / hide / 角色定义）",
        "  · 变量名列：只填变量（变量定义 / 赋值 / 增减 / 开关 / 条件比较）",
        "  · 图片列：下拉可选择已定义的 image 名或图片文件路径",
        "",
        "【结构化指令类型】",
        "  定义类：",
        "    角色定义（定义角色）  — 角色名列填名字，对话文本列填 Character(...)",
        "    变量定义（定义变量）  — 变量名列填名字，对话文本列填初始值",
        "    图片定义（定义图片）  — 图片/背景列填路径，变量名列填标识名",
        "  操作类：",
        "    变量赋值（变量赋值）  — 变量名列填变量名，对话文本列填新值",
        "    变量增减（变量增减）  — 变量名列填变量名，对话文本列填增量（如 1 或 -3）",
        "    变量开关（变量开关）  — 变量名列填变量名，对话文本列下拉选 true/false",
        "  条件类：",
        "    变量=（等于）         — 变量名列填变量名，对话文本列填纯数值",
        "    变量≠（不等于）       — 同理",
        "    变量>（大于）         — 同理",
        "    变量≥（大于等于）     — 同理",
        "    变量<（小于）         — 同理",
        "    变量≤（小于等于）     — 同理",
        "    else（否则）          — 传统 else",
        "  复杂条件仍可用传统 if/elif（对话文本列写完整表达式）",
        "",
        "【下拉菜单】",
        "  · 角色名列：自动扫描 game/*.rpy 中的 define + 表中定义的",
        "  · 变量名列：自动扫描 game/*.rpy 中的 default + 表中定义的",
        "  · 图片列： 自动扫描 image 定义名 + game/ 下所有图片文件",
        "  · 指令类型：固定下拉列表",
        "  · 变量开关：固定 true / false 下拉",
        "",
        "【智能填充】",
        "  · 角色名/变量名列留空 → 自动沿用上一行的值",
        "  · 指令类型留空 + 有角色名 → 自动当 dialogue",
        "  · 指令类型留空 + 有文本 + 无角色 → 自动当 旁白",
        "  · 角色名填「旁白」→ 自动切换 narrator",
        "",
        "【中文术语】（位置/特效列可用中文）",
        "  转场：溶解 褪色 闪白 像素化 横向振动 纵向振动 百叶窗 网格覆盖 擦除 滑入 滑出 推出",
        "  位置：左边 右边 中间 真中 左外 右外",
        "  空格分隔可组合：右边 溶解  →  at right with dissolve",
        "",
        "【注意事项】",
        "  · 空行分隔场景区块、结束 if/menu 块",
        "  · 图片路径相对 Ren'Py 的 game/ 目录",
        "  · 函数调用（Transform/Solid）直接写，不加引号",
        "  · 对话中引用变量用 [变量名]",
        "  · menu_option 的跳转目标支持 jump/call/return/$",
    ]
    for row_idx, text in enumerate(help_lines, start=1):
        cell = ws_help.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name="微软雅黑", size=10)
        if text.startswith("═══") or text.startswith("【"):
            cell.font = Font(name="微软雅黑", bold=True, size=11)
    ws_help.column_dimensions["A"].width = 80

    _safe_save_workbook(wb, output_path, "模板")


def generate_blank_template(output_path: str):
    """生成纯空白模板（只有表头+下拉验证，无示例数据，无使用说明）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "script"

    scanned_characters, scanned_variables, scanned_images, scanned_files = scan_existing_scripts()
    _setup_sheet_header_and_dropdowns(
        ws, scanned_characters, scanned_variables, scanned_images, scanned_files
    )

    _safe_save_workbook(wb, output_path, "空白模板")


# ── 转换逻辑 ──────────────────────────────────────────────


def _get_desktop_documents() -> list:
    """获取桌面和文档目录候选（兼容中英文 Windows），返回 [(路径, 友好名称), ...]"""
    from pathlib import Path as _Path

    home = _Path.home()
    seen = set()
    result = []
    for name, label in [
        ("Desktop", "桌面"), ("desktop", "桌面"), ("桌面", "桌面"),
        ("Documents", "文档"), ("documents", "文档"), ("文档", "文档"),
        ("My Documents", "文档"), ("我的文档", "文档"),
    ]:
        p = home / name
        if p.is_dir() and p not in seen:
            seen.add(p)
            result.append((p, label))
    return result


def _try_copy_to_fallbacks(src_data: bytes, stem: str, ext: str, target_parent, label: str) -> str:
    """将 bytes 内容依次尝试写入目标目录或回退目录。成功返回路径，失败返回空字符串。"""
    from pathlib import Path as _Path

    fallback_dirs = [(target_parent, "当前目录")] + _get_desktop_documents()

    for dest_dir, dir_label in fallback_dirs:
        dest = dest_dir / f"{stem}{ext}"
        try:
            dest.write_bytes(src_data)
            if dest_dir != target_parent:
                print(f"\n⚠️  当前目录写入被拦截，已自动保存到{dir_label}：")
            else:
                print(f"[OK] {label} saved:")
            print(f"   {dest}")
            return str(dest)
        except (PermissionError, OSError):
            for i in range(1, 100):
                alt = dest_dir / f"{stem}_{i}{ext}"
                try:
                    alt.write_bytes(src_data)
                    if dest_dir != target_parent:
                        print(f"\n[WARN] Can't write to target dir, saved to {dir_label}:")
                    else:
                        print(f"[OK] {label} saved:")
                    print(f"   {alt}")
                    return str(alt)
                except (PermissionError, OSError):
                    continue
            continue
    return ""


def _safe_save_workbook(wb, output_path: str, label: str):
    """
    安全保存 Excel 工作簿：用 BytesIO 绕过 openpyxl 的内部重命名，
    直接写字节到目标路径。被拦截时依次尝试桌面、文档。返回实际保存路径。
    """
    from io import BytesIO
    from pathlib import Path as _Path

    target = _Path(output_path).resolve()
    stem = target.stem
    ext = target.suffix

    buffer = BytesIO()
    try:
        wb.save(buffer)
    except Exception:
        try:
            wb.save(str(target))
            print(f"[OK] {label} saved:")
            print(f"   {target}")
            return str(target)
        except (PermissionError, OSError):
            buffer = BytesIO()
            wb.save(buffer)

    data = buffer.getvalue()

    last_error = None
    for attempt in range(3):
        try:
            target.write_bytes(data)
            print(f"[OK] {label} saved:")
            print(f"   {target}")
            return str(target)
        except (PermissionError, OSError) as e:
            last_error = e
            if attempt < 2:
                import time
                time.sleep(0.3)
            continue

    print(f"\n⚠️  无法写入目标目录：{target.parent}")
    print(f"   原因：{last_error}")
    result = _try_copy_to_fallbacks(data, stem, ext, target.parent, label)
    if result:
        return result

    import tempfile
    tmp_dir = _Path(tempfile.gettempdir())
    tmp_path = tmp_dir / f"{stem}{ext}"
    tmp_path.write_bytes(data)
    print(f"\n⚠️  无法写入任何目录（Defender 拦截？），文件保留在临时目录：")
    print(f"   {tmp_path}")
    return str(tmp_path)


def _safe_write_text(output_path: str, content: str):
    """
    安全写入文本文件：优先直接写入目标路径，被拦截时依次尝试桌面、文档。
    返回实际保存的路径。
    """
    from pathlib import Path as _Path

    target = _Path(output_path).resolve()
    stem = target.stem
    ext = target.suffix
    data = content.encode("utf-8")

    for attempt in range(3):
        try:
            target.write_bytes(data)
            return str(target)
        except (PermissionError, OSError):
            if attempt < 2:
                import time
                time.sleep(0.3)
            continue

    result = _try_copy_to_fallbacks(data, stem, ext, target.parent, "文件")
    if result:
        return result

    import tempfile
    tmp_dir = _Path(tempfile.gettempdir())
    tmp_path = tmp_dir / f"{stem}{ext}"
    tmp_path.write_bytes(data)
    print(f"\n⚠️  无法写入任何目录（Defender 拦截？），文件保留在临时目录：")
    print(f"   {tmp_path}")
    return str(tmp_path)


# ── 转换逻辑（续）──────────────────────────────────────────


def _val(row, col_idx: int, default=""):
    """安全获取单元格值"""
    if col_idx >= len(row):
        return default
    v = row[col_idx]
    return str(v).strip() if v is not None else default


def _escape_rpy(text: str) -> str:
    """转义 Ren'Py 字符串中的特殊字符（仅转义双引号）"""
    return text.replace('"', '\\"')


def _parse_cmd(raw: str) -> str:
    """从指令单元格中提取英文指令名（兼容 'jump（跳转）' 格式）"""
    return raw.split("（")[0].strip().lower()


def _is_structured_variable_cmd(cmd: str) -> bool:
    """判断是否为结构化变量类指令（定义/赋值/增减/开关/条件比较）"""
    return cmd in (
        "define_variable", "variable_set", "variable_change", "variable_toggle",
        "variable_eq", "variable_ne", "variable_gt", "variable_ge",
        "variable_lt", "variable_le",
    )


# ── 中文术语映射 ──────────────────────────────────────────

TRANSITION_MAP = {
    "溶解": "dissolve", "褪色": "fade", "闪白": 'Fade(0.1,0.0,0.5,color="#FFFFFF")',
    "像素化": "pixellate", "横向振动": "hpunch", "纵向振动": "vpunch",
    "百叶窗": "blinds", "网格覆盖": "squares", "擦除": "wipeleft",
    "滑入": "slideleft", "滑出": "slideawayleft", "推出": "pushright",
}

POSITION_MAP = {
    "左边": "left", "右边": "right", "中间": "center",
    "真中": "truecenter", "左外": "offscreenleft", "右外": "offscreenright",
}

AUDIO_CMD_MAP = {
    "停止": "stop", "播放": "play", "循环": "loop",
}


def _translate_effect(effect: str) -> str:
    """翻译中文特效/位置为 Ren'Py 英文"""
    if not effect:
        return ""
    parts = effect.split()
    translated = []
    for p in parts:
        if p in TRANSITION_MAP:
            translated.append(TRANSITION_MAP[p])
        elif p in POSITION_MAP:
            translated.append(POSITION_MAP[p])
        else:
            translated.append(p)
    return " ".join(translated)


def _trim_role_name(name: str) -> str:
    """修剪角色名首尾空白，收集警告"""
    trimmed = name.strip()
    if trimmed != name and name.strip():
        return trimmed
    return name


# ── 表格校验 ──────────────────────────────────────────────

def check_excel(input_path: str) -> list:
    """
    转换前校验 Excel，返回警告列表 [(row_num, severity, message), ...]。
    severity: "error" 可能导致语法错误, "warn" 建议检查
    """
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=True)
    issues = []

    all_labels = set()
    # 收集结构化变量名
    all_variables = set()

    valid_cmds = {
        "label", "scene", "show", "hide", "dialogue", "narrator",
        "menu", "menu_option", "jump", "call", "return",
        "play_music", "stop_music", "queue_music", "play_sound", "stop_sound",
        "voice", "pause", "player_input", "window",
        "define_character", "define_variable", "define_image",
        "variable_set", "variable_change", "variable_toggle",
        "variable_eq", "variable_ne", "variable_gt", "variable_ge",
        "variable_lt", "variable_le",
        "$", "if", "elif", "else",
        "default", "image",  # 向后兼容
    }

    for sheet in wb.worksheets:
        if sheet.sheet_state == "hidden" or sheet.title.startswith("_dd"):
            continue
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue

        # 第一遍：收集 label 和变量名
        for row_idx, raw_row in enumerate(rows):
            row = [str(c).strip() if c is not None else "" for c in raw_row]
            if all(c == "" for c in row):
                continue
            cmd = _parse_cmd(_val(row, COL_CMD))
            scene_label = _val(row, COL_SCENE_LABEL)
            dialogue = _val(row, COL_DIALOGUE)
            variable = _val(row, COL_VARIABLE)

            if cmd == "label":
                name = scene_label if scene_label else dialogue
                if name and name in all_labels:
                    issues.append((row_idx + 2, "error", f"重复的标签名：{name}"))
                elif name:
                    all_labels.add(name)

            if _is_structured_variable_cmd(cmd) and variable:
                all_variables.add(variable)

        # 第二遍：逐行检查
        in_menu = False
        menu_has_options = False
        in_if = False
        if_has_body = False
        last_role = ""
        last_variable = ""

        for row_idx, raw_row in enumerate(rows):
            excel_row = row_idx + 2
            row = [str(c).strip() if c is not None else "" for c in raw_row]
            if all(c == "" for c in row):
                if in_menu and not menu_has_options:
                    issues.append((menu_start_row, "error", "菜单没有任何选项"))
                in_menu = False
                menu_has_options = False
                if in_if and not if_has_body:
                    issues.append((if_start_row, "error", "if/elif/else 没有后续指令"))
                in_if = False
                if_has_body = False
                continue

            raw_cmd = _val(row, COL_CMD)
            cmd = _parse_cmd(raw_cmd)
            scene_label = _val(row, COL_SCENE_LABEL)
            image_path = _val(row, COL_IMAGE)
            character = _val(row, COL_CHARACTER)
            variable = _val(row, COL_VARIABLE)
            dialogue = _val(row, COL_DIALOGUE)
            option_text = _val(row, COL_OPTION)
            jump_target = _val(row, COL_JUMP)
            audio = _val(row, COL_AUDIO)
            effect = _val(row, COL_EFFECT)

            # 前向填充
            if character.strip():
                last_role = character.strip()
            role = last_role
            if variable.strip():
                last_variable = variable.strip()
            var_name = last_variable

            # 未知指令
            if raw_cmd.strip() and cmd not in valid_cmds:
                issues.append((excel_row, "warn", f"未知指令类型：{raw_cmd}"))

            # label 无名称
            if cmd == "label":
                name = scene_label if scene_label else dialogue
                if not name:
                    issues.append((excel_row, "error", "label 没有场景标签名"))

            # scene 无背景
            if cmd == "scene":
                if not image_path and not effect:
                    issues.append((excel_row, "warn", "scene 没有指定背景图"))

            # show 无图片
            if cmd == "show":
                if not image_path and not character:
                    issues.append((excel_row, "warn", "show 没有指定图片或角色"))

            # dialogue 无文本
            if cmd == "dialogue":
                if not role or role == "narrator":
                    issues.append((excel_row, "warn", "dialogue 没有角色名"))
                if not dialogue:
                    issues.append((excel_row, "warn", "dialogue 没有对话文本"))

            if cmd == "narrator":
                if not dialogue:
                    issues.append((excel_row, "warn", "narrator 没有文本"))

            # 结构化变量检查
            if cmd == "define_variable":
                if not var_name:
                    issues.append((excel_row, "warn", "变量定义缺少变量名"))
                if not dialogue:
                    issues.append((excel_row, "warn", "变量定义缺少初始值"))

            if cmd == "variable_set":
                if not var_name:
                    issues.append((excel_row, "warn", "变量赋值缺少变量名"))
                if not dialogue:
                    issues.append((excel_row, "warn", "变量赋值缺少值"))

            if cmd == "variable_change":
                if not var_name:
                    issues.append((excel_row, "warn", "变量增减缺少变量名"))
                if not dialogue:
                    issues.append((excel_row, "warn", "变量增减缺少增量值"))

            if cmd == "variable_toggle":
                if not var_name:
                    issues.append((excel_row, "warn", "变量开关缺少变量名"))
                if dialogue.lower() not in ("true", "false"):
                    issues.append((excel_row, "warn", "变量开关值应为 true 或 false"))

            # 结构化条件检查
            if cmd in VARIABLE_OP_MAP:
                if not var_name:
                    issues.append((excel_row, "warn", f"{cmd} 缺少变量名"))
                if not dialogue:
                    issues.append((excel_row, "warn", f"{cmd} 缺少比较值"))
                lg = _val(row, COL_CONNECT).strip().lower()
                if lg and lg not in ("and", "or"):
                    issues.append((excel_row, "warn", f"逻辑连接值无效：'{lg}'，应使用 and / or / 空"))

            # menu 检测
            if cmd == "menu":
                if in_menu:
                    if not menu_has_options:
                        issues.append((menu_start_row, "error", "菜单没有任何选项"))
                in_menu = True
                menu_has_options = False
                menu_start_row = excel_row

            if cmd == "menu_option":
                if not in_menu:
                    issues.append((excel_row, "error", "menu_option 不在 menu 内"))
                else:
                    menu_has_options = True
                if not option_text:
                    issues.append((excel_row, "warn", "menu_option 没有选项文本"))
                if not jump_target:
                    issues.append((excel_row, "warn", "menu_option 没有跳转目标"))

            if cmd not in ("menu", "menu_option") and in_menu:
                if not menu_has_options:
                    issues.append((menu_start_row, "error", "菜单没有任何选项"))
                in_menu = False
                menu_has_options = False

            # if/elif/else 检测（传统条件 和 结构化条件）
            is_condition = cmd in VARIABLE_OP_MAP or cmd in ("if", "elif", "else")
            if is_condition:
                if in_if and not if_has_body:
                    issues.append((if_start_row, "error", "if 块没有后续指令"))
                in_if = True
                if_has_body = False
                if_start_row = excel_row
                if cmd in ("if", "elif") and not dialogue:
                    issues.append((excel_row, "warn", "if/elif 没有条件表达式"))
            elif in_if:
                if_has_body = True

            # jump/call 目标检查
            if cmd in ("jump", "call"):
                if not jump_target:
                    issues.append((excel_row, "warn", f"{cmd} 没有跳转目标"))
                elif jump_target not in all_labels:
                    issues.append((excel_row, "warn", f"跳转目标 '{jump_target}' 未定义（可能在其他 Sheet）"))

            # 音频检查
            if cmd in ("play_music", "play_sound", "queue_music", "voice"):
                if not audio:
                    issues.append((excel_row, "warn", f"{cmd} 没有音频路径"))

            # player_input 检查
            if cmd == "player_input":
                if not character:
                    issues.append((excel_row, "warn", "player_input 没有变量名"))

    return issues


def _print_check_result(issues: list, max_show: int = 20):
    """打印校验结果"""
    if not issues:
        print("  Check: No issues found.")
        return

    errors = [i for i in issues if i[1] == "error"]
    warns = [i for i in issues if i[1] == "warn"]
    print(f"  Check: {len(errors)} error(s), {len(warns)} warning(s)")

    shown = 0
    for row, sev, msg in errors + warns:
        if shown >= max_show:
            print(f"  ... and {len(issues) - max_show} more")
            break
        tag = "ERROR" if sev == "error" else "WARN"
        print(f"  [{tag}] 行 {row}: {msg}")
        shown += 1


def convert_excel_to_rpy(input_path: str, output_path: str):
    """将 Excel 转为 .rpy 脚本（支持多 Sheet）"""
    wb = load_workbook(input_path, data_only=True)

    warnings = []

    def _warn(msg: str, row_num: int = 0):
        sheet_name = wb.active.title if wb.active else "?"
        warnings.append(f"[{sheet_name} 行{row_num + 1}] {msg}")

    all_sheet_lines = []
    all_header_lines = []
    defined_characters = set()
    role_registry = set()
    # 变量注册表（自动收集）
    variable_registry = set()

    for sheet in wb.worksheets:
        if sheet.sheet_state == "hidden" or sheet.title.startswith("_dd"):
            continue
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue

        lines = []
        header_lines = []
        in_menu = False
        in_if_body = False
        base_indent = "    "

        current_role = ""
        current_variable = ""
        current_characters = []
        scene_sheet_label = ""
        # 复合条件链状态
        pending_chain_parts = []   # 累积的条件片段
        prev_connector = ""        # 上一行的逻辑连接符

        def _indent():
            return base_indent + ("    " if in_if_body else "")

        def _end_if_body():
            nonlocal in_if_body
            in_if_body = False

        for row_idx, raw_row in enumerate(rows):
            row = [str(c).strip() if c is not None else "" for c in raw_row]

            if all(c == "" for c in row):
                _end_if_body()
                if in_menu:
                    in_menu = False
                # 空行 → 结束未完成的复合条件链
                if pending_chain_parts:
                    lines.append(f"{base_indent}{' '.join(pending_chain_parts)}:")
                    pending_chain_parts = []
                    prev_connector = ""
                    in_if_body = True
                continue

            scene_label = _val(row, COL_SCENE_LABEL)
            raw_cmd = _val(row, COL_CMD)
            cmd = _parse_cmd(raw_cmd)
            image_path = _val(row, COL_IMAGE)
            character = _val(row, COL_CHARACTER)
            variable = _val(row, COL_VARIABLE)
            logic_connect = _val(row, COL_CONNECT).strip().lower()
            dialogue = _val(row, COL_DIALOGUE)
            option_text = _val(row, COL_OPTION)
            jump_target = _val(row, COL_JUMP)
            audio = _val(row, COL_AUDIO)
            effect = _translate_effect(_val(row, COL_EFFECT))
            notes = _val(row, COL_NOTES)

            # 向后兼容旧指令
            if cmd in OLD_CMD_ALIASES:
                cmd = OLD_CMD_ALIASES[cmd]

            # 角色名前向填充
            if character.strip():
                trimmed = _trim_role_name(character)
                if trimmed != character and character.strip():
                    _warn(f"角色名首尾含空白，已修剪：'{character}' → '{trimmed}'", row_idx)
                current_role = trimmed
            role = current_role

            # 变量名前向填充
            if variable.strip():
                current_variable = variable.strip()
            var_name = current_variable

            # 旁白自动检测
            if role == "旁白":
                role = "narrator"

            # 空指令 + 有角色名 → 自动推断为 dialogue
            if not raw_cmd.strip() and role and role != "narrator":
                cmd = "dialogue"
            # 空指令 + 有对话文本 + 无角色 → 旁白
            if not raw_cmd.strip() and dialogue and not role:
                cmd = "narrator"

            # ── 记录角色和变量 ──
            if role and role != "narrator":
                role_registry.add(role)
            if _is_structured_variable_cmd(cmd) and var_name:
                variable_registry.add(var_name)

            # ── define_character ──
            if cmd == "define_character":
                name = character if character else "unknown"
                defined_characters.add(name)
                if dialogue:
                    line = f"define {name} = {dialogue}"
                else:
                    line = f'define {name} = Character("{name}")'
                header_lines.append(line)
                continue

            # ── define_variable ──
            if cmd == "define_variable":
                name = var_name if var_name else "unknown"
                val = dialogue if dialogue else '""'
                header_lines.append(f"default {name} = {val}")
                continue

            # ── define_image ──
            if cmd == "define_image":
                img_name = var_name if var_name else "unknown"
                img_path = image_path if image_path else ""
                if "(" in img_path:
                    header_lines.append(f'image {img_name} = {img_path}')
                elif img_path:
                    header_lines.append(f'image {img_name} = "{_escape_rpy(img_path)}"')
                else:
                    header_lines.append(f'image {img_name}')
                continue

            # ── label ──
            if cmd == "label":
                _end_if_body()
                in_menu = False
                label_name = scene_label if scene_label else dialogue
                if not label_name:
                    continue
                lines.append(f"label {label_name}:")
                continue

            # ── scene ──
            if cmd == "scene":
                img = _escape_rpy(image_path) if image_path else ""
                fx = f" {effect}" if effect else ""
                if img:
                    lines.append(f"{_indent()}scene {img}{fx}")
                elif effect:
                    lines.append(f"{_indent()}scene{fx}")
                continue

            # ── show ──
            if cmd == "show":
                char = character if character else ""
                img = _escape_rpy(image_path) if image_path else ""
                fx = f" {effect}" if effect else ""
                if char and img:
                    lines.append(f"{_indent()}show {char} {img}{fx}")
                elif char:
                    lines.append(f"{_indent()}show {char}{fx}")
                elif img:
                    lines.append(f"{_indent()}show {img}{fx}")
                continue

            # ── hide ──
            if cmd == "hide":
                char = character if character else ""
                fx = f" {effect}" if effect else ""
                if char:
                    lines.append(f"{_indent()}hide {char}{fx}")
                continue

            # ── dialogue ──
            if cmd == "dialogue":
                char = character if character else ""
                text = _escape_rpy(dialogue) if dialogue else ""
                if char and text:
                    lines.append(f'{_indent()}{char} "{text}"')
                continue

            # ── narrator ──
            if cmd == "narrator":
                text = _escape_rpy(dialogue) if dialogue else ""
                if text:
                    if notes and "centered" in notes.lower():
                        lines.append(f'{_indent()}centered "{text}"')
                    elif notes and "explicit" in notes.lower():
                        lines.append(f'{_indent()}narrator "{text}"')
                    else:
                        lines.append(f'{_indent()}"{text}"')
                continue

            # ── menu ──
            if cmd == "menu":
                _end_if_body()
                lines.append(f"{_indent()}menu:")
                in_menu = True
                continue

            # ── menu_option ──
            if cmd == "menu_option":
                opt = _escape_rpy(option_text) if option_text else ""
                target = jump_target if jump_target else ""
                if opt and target:
                    lines.append(f'{_indent()}    "{opt}":')
                    if target.startswith("jump "):
                        lines.append(f"{_indent()}        jump {target[5:]}")
                    elif target in ("return",) or target.startswith("call ") or target.startswith("$ "):
                        lines.append(f"{_indent()}        {target}")
                    else:
                        lines.append(f"{_indent()}        jump {target}")
                elif opt:
                    lines.append(f'{_indent()}    "{opt}":')
                    lines.append(f"{_indent()}        pass")
                continue

            if cmd != "menu_option" and in_menu:
                in_menu = False

            # ── jump ──
            if cmd == "jump":
                target = jump_target if jump_target else ""
                if target:
                    lines.append(f"{_indent()}jump {target}")
                continue

            # ── call ──
            if cmd == "call":
                target = jump_target if jump_target else ""
                if target:
                    lines.append(f"{_indent()}call {target}")
                continue

            # ── return ──
            if cmd == "return":
                lines.append(f"{_indent()}return")
                continue

            # ── variable_set（结构化变量赋值）──
            if cmd == "variable_set":
                name = var_name if var_name else ""
                val = dialogue if dialogue else ""
                if name and val:
                    lines.append(f"{_indent()}$ {name} = {val}")
                continue

            # ── variable_change（结构化变量增减）──
            if cmd == "variable_change":
                name = var_name if var_name else ""
                delta = dialogue if dialogue else "0"
                if name:
                    if delta.startswith("-"):
                        lines.append(f"{_indent()}$ {name} -= {delta[1:]}")
                    elif delta.startswith("+"):
                        lines.append(f"{_indent()}$ {name} += {delta[1:]}")
                    else:
                        lines.append(f"{_indent()}$ {name} += {delta}")
                continue

            # ── variable_toggle（结构化变量开关）──
            if cmd == "variable_toggle":
                name = var_name if var_name else ""
                val = dialogue.strip().lower() if dialogue else "true"
                if name:
                    val_bool = "True" if val == "true" else "False"
                    lines.append(f"{_indent()}$ {name} = {val_bool}")
                continue

            # ── 结构化条件判断（支持复合条件链）──
            if cmd in VARIABLE_OP_MAP:
                name = var_name if var_name else ""
                val = dialogue if dialogue else "0"
                op = VARIABLE_OP_MAP[cmd]

                if logic_connect:
                    # 复合条件的一部分
                    if not pending_chain_parts:
                        _end_if_body()
                        pending_chain_parts.append(f"if {name} {op} {val}")
                    else:
                        pending_chain_parts.append(f"{prev_connector} {name} {op} {val}")
                    prev_connector = logic_connect
                else:
                    # 逻辑连接为空
                    if pending_chain_parts:
                        pending_chain_parts.append(f"{prev_connector} {name} {op} {val}")
                        lines.append(f"{base_indent}{' '.join(pending_chain_parts)}:")
                        pending_chain_parts = []
                        prev_connector = ""
                        in_if_body = True
                    else:
                        _end_if_body()
                        lines.append(f"{_indent()}if {name} {op} {val}:")
                        in_if_body = True
                continue

            # 其他指令出现 → 结束未完成的复合条件链
            if pending_chain_parts:
                lines.append(f"{base_indent}{' '.join(pending_chain_parts)}:")
                pending_chain_parts = []
                prev_connector = ""
                in_if_body = True

            # ── $（通用设置变量）──
            if cmd == "$":
                code = dialogue if dialogue else ""
                if code:
                    lines.append(f"{_indent()}$ {code}")
                continue

            # ── if ──
            if cmd == "if":
                _end_if_body()
                cond = dialogue if dialogue else "True"
                lines.append(f"{_indent()}if {cond}:")
                in_if_body = True
                continue

            # ── elif ──
            if cmd == "elif":
                _end_if_body()
                cond = dialogue if dialogue else "True"
                lines.append(f"{_indent()}elif {cond}:")
                in_if_body = True
                continue

            # ── else ──
            if cmd == "else":
                _end_if_body()
                lines.append(f"{_indent()}else:")
                in_if_body = True
                continue

            # ── play_music ──
            if cmd == "play_music":
                aud = _escape_rpy(audio) if audio else ""
                if aud:
                    lines.append(f'{_indent()}play music "{aud}"')
                continue

            # ── stop_music ──
            if cmd == "stop_music":
                lines.append(f"{_indent()}stop music {effect}" if effect else f"{_indent()}stop music")
                continue

            # ── play_sound ──
            if cmd == "play_sound":
                aud = _escape_rpy(audio) if audio else ""
                if aud:
                    lines.append(f'{_indent()}play sound "{aud}"')
                continue

            # ── stop_sound ──
            if cmd == "stop_sound":
                lines.append(f"{_indent()}stop sound")
                continue

            # ── queue_music ──
            if cmd == "queue_music":
                aud = _escape_rpy(audio) if audio else ""
                if aud:
                    lines.append(f'{_indent()}queue music "{aud}"')
                continue

            # ── voice ──
            if cmd == "voice":
                aud = _escape_rpy(audio) if audio else ""
                if aud:
                    lines.append(f'{_indent()}voice "{aud}"')
                continue

            # ── window ──
            if cmd == "window":
                action = _escape_rpy(dialogue) if dialogue else "show"
                if action in ("show", "hide", "auto"):
                    lines.append(f"{_indent()}window {action}")
                else:
                    lines.append(f"{_indent()}window show")
                continue

            # ── pause ──
            if cmd == "pause":
                dur = dialogue if dialogue else "1.0"
                try:
                    float(dur)
                except ValueError:
                    dur = "1.0"
                lines.append(f"{_indent()}pause {dur}")
                continue

            # ── player_input（玩家输入）──
            if cmd == "player_input":
                var_name_input = character if character else "input_result"
                prompt = dialogue if dialogue else "请输入："
                default_val = effect if effect else ""
                prompt_escaped = _escape_rpy(prompt)
                if default_val:
                    lines.append(f'{_indent()}$ {var_name_input} = renpy.input("{prompt_escaped}").strip() or "{_escape_rpy(default_val)}"')
                else:
                    lines.append(f'{_indent()}$ {var_name_input} = renpy.input("{prompt_escaped}").strip()')
                continue

            # show 指令中的图片名注册
            if cmd == "show" and image_path:
                img_name = image_path.split()[0] if image_path else ""
                if img_name and not img_name.startswith("bg "):
                    role_registry.add(img_name)

        all_sheet_lines.append((scene_sheet_label or sheet.title, lines, header_lines))

    # ── 自动生成未定义的 Character ──
    auto_defines = role_registry - defined_characters - {"narrator", ""}
    for name in sorted(auto_defines):
        all_header_lines.insert(0, f'define {name} = Character("{name}")')

    # ── 组装输出 ──
    first_sheet_headers = all_sheet_lines[0][2] if all_sheet_lines else []
    output_parts = ["# Generated by excel_to_rpy.py",
                    "# 源文件: " + os.path.basename(input_path),
                    ""]
    output_parts.extend(all_header_lines)
    output_parts.extend(first_sheet_headers)
    output_parts.append("")

    total_lines = 0
    first = True
    for label, lines, _headers in all_sheet_lines:
        if first:
            first = False
        else:
            output_parts.append("")
            has_label = any(l.strip().startswith("label ") for l in lines)
            if not has_label:
                output_parts.append(f"label {label}:")
        output_parts.extend(lines)
        total_lines += len(lines)

    output = "\n".join(output_parts) + "\n"
    saved_path = _safe_write_text(output_path, output)

    print(f"Complete: {saved_path}")
    print(f"   {total_lines} lines / {len(all_sheet_lines)} sheet(s)")

    if warnings:
        print(f"\n[WARN] {len(warnings)} issue(s):")
        for w in warnings:
            print(f"  {w}")

    if auto_defines:
        print(f"\n[INFO] Auto-defined {len(auto_defines)} character(s):")
        for n in sorted(auto_defines):
            print(f"  define {n} = Character(\"{n}\")")

    # ── 自动追加定义到 defines.rpy ──
    try:
        append_defines_from_excel(input_path)
    except Exception as e:
        print(f"[INFO] 定义追加跳过: {e}")


# ── 定义管理模块 ──────────────────────────────────────────

def _find_defines_rpy():
    """查找 defines.rpy 路径（位于 game/ 目录下）。返回 Path 或 None。"""
    root = _find_project_root()
    game_path = root / "game"
    if game_path.is_dir():
        return game_path / "defines.rpy"
    return None


def _find_game_dir():
    """查找 game/ 目录。返回 Path 或 None。"""
    root = _find_project_root()
    game_path = root / "game"
    return game_path if game_path.is_dir() else None


def scan_excel_for_defines(excel_path):
    """从 Excel 中提取所有 define_character / define_variable / define_image 定义。

    返回: [{type, name, line, source}, ...]
      type: "character" | "variable" | "image"
    """
    from pathlib import Path as _Path

    if not os.path.isfile(excel_path):
        return []

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception:
        return []

    defines = []
    source_label = f"excel:{_Path(excel_path).name}"

    for sheet in wb.worksheets:
        if sheet.sheet_state == "hidden" or sheet.title.startswith("_dd"):
            continue
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue

        current_role = ""
        current_variable = ""

        for raw_row in rows:
            row_data = [str(c).strip() if c is not None else "" for c in raw_row]
            if all(c == "" for c in row_data):
                current_role = ""
                current_variable = ""
                continue

            raw_cmd = _val(row_data, COL_CMD)
            cmd = _parse_cmd(raw_cmd)
            image_path = _val(row_data, COL_IMAGE)
            character = _val(row_data, COL_CHARACTER)
            variable = _val(row_data, COL_VARIABLE)
            dialogue = _val(row_data, COL_DIALOGUE)

            if cmd in OLD_CMD_ALIASES:
                cmd = OLD_CMD_ALIASES[cmd]

            if character.strip():
                current_role = character.strip()
            if variable.strip():
                current_variable = variable.strip()

            if cmd == "define_character":
                name = current_role if current_role else "unknown"
                if dialogue:
                    line = f"define {name} = {dialogue}"
                else:
                    line = f'define {name} = Character("{name}")'
                defines.append({
                    "type": "character",
                    "name": name,
                    "line": line,
                    "source": source_label,
                })

            elif cmd == "define_variable":
                name = current_variable if current_variable else "unknown"
                val = dialogue if dialogue else '""'
                defines.append({
                    "type": "variable",
                    "name": name,
                    "line": f"default {name} = {val}",
                    "source": source_label,
                })

            elif cmd == "define_image":
                img_name = current_variable if current_variable else "unknown"
                img_path = image_path if image_path else ""
                if "(" in img_path:
                    line = f"image {img_name} = {img_path}"
                elif img_path:
                    line = f'image {img_name} = "{_escape_rpy(img_path)}"'
                else:
                    line = f"image {img_name}"
                defines.append({
                    "type": "image",
                    "name": img_name,
                    "line": line,
                    "source": source_label,
                })

    return defines


# Ren'Py 框架自带的系统文件，不应纳入定义管理
_RENPY_SYSTEM_FILES = {
    "screens.rpy", "gui.rpy", "options.rpy", "script_version.rpy",
    "navigation.rpy",
}


def scan_rpy_for_defines(rpy_path):
    """从 .rpy 文件中提取所有顶层 define/default/image 定义。
    排除 defines.rpy 自身和 Ren'Py 系统文件。

    返回: [{type, name, line, source}, ...]
    """
    path_obj = Path(rpy_path)
    if path_obj.name.lower() == "defines.rpy":
        return []
    if path_obj.name.lower() in _RENPY_SYSTEM_FILES:
        return []

    try:
        content = path_obj.read_text(encoding="utf-8-sig")
    except Exception:
        return []

    defines = []
    source_label = f"rpy:{path_obj.name}"

    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("#") or not stripped:
            continue

        m = re.match(r'define\s+(\w+)\s*=\s*Character\b', stripped)
        if m:
            defines.append({
                "type": "character",
                "name": m.group(1),
                "line": stripped,
                "source": source_label,
            })
            continue

        m = re.match(r'default\s+(\w+)\s*=\s*(.+)', stripped)
        if m:
            defines.append({
                "type": "variable",
                "name": m.group(1),
                "line": stripped,
                "source": source_label,
            })
            continue

        m = re.match(r'image\s+(.+?)\s*=\s*(.+)', stripped)
        if m:
            defines.append({
                "type": "image",
                "name": m.group(1).strip(),
                "line": stripped,
                "source": source_label,
            })

    return defines


def parse_existing_defines(defines_path):
    """解析现有的 defines.rpy，返回 {name: line} 映射。"""
    if not os.path.isfile(defines_path):
        return {}

    try:
        content = Path(defines_path).read_text(encoding="utf-8-sig")
    except Exception:
        return {}

    existing = {}
    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("#") or stripped.startswith("=") or not stripped:
            continue

        m = re.match(r'define\s+(\w+)\s*=\s*(.+)', stripped)
        if m:
            existing[m.group(1)] = stripped
            continue

        m = re.match(r'default\s+(\w+)\s*=\s*(.+)', stripped)
        if m:
            existing[m.group(1)] = stripped
            continue

        m = re.match(r'image\s+(.+?)\s*=\s*(.+)', stripped)
        if m:
            existing[m.group(1).strip()] = stripped

    return existing


def _match_def_to_character(def_name, def_type, characters):
    """将定义匹配到角色名。返回角色名或 None（系统/全局）。
    匹配策略：在定义名中用下划线和空格分割后，查找与角色名相同的片段。
    最长匹配优先；角色本身的 define 直接匹配自身。
    """
    if def_type == "character":
        return def_name

    def_parts = re.split(r'[\s_]+', def_name.lower())
    best = None
    best_len = 0

    for char in characters:
        char_lower = char.lower()
        for part in def_parts:
            if part == char_lower and len(char) > best_len:
                best = char
                best_len = len(char)

    return best


# 虚拟角色前缀：变量/图片名前缀 → 分组键 + 显示名
_VIRTUAL_CHAR_PREFIXES = [
    (["player_", "pc_", "mc_", "主角_", "主角"], "_player", "玩家"),
    (["bg_", "bg ", "background_", "background "], "_bg", "背景"),
]


def _match_def_to_virtual(def_name):
    """匹配虚拟角色前缀（玩家、背景等）。返回 (group_key, display_name) 或 None。
    支持前缀（player_name → player）和后缀（score_player → player）两种命名风格。
    """
    name_lower = def_name.lower()
    tokens = name_lower.replace(" ", "_").split("_")
    for prefixes, key, label in _VIRTUAL_CHAR_PREFIXES:
        for prefix in prefixes:
            if name_lower.startswith(prefix.lower()):
                return (key, label)
            clean = prefix.lower().rstrip("_ ")
            if clean and clean in tokens:
                return (key, label)
    return None


# 分类前缀：名称首词命中这些前缀时，优先归入对应虚拟分组（不参与角色匹配）
_CATEGORY_FIRST_TOKENS = {
    "bg": ("_bg", "背景"),
}


def _group_defines_by_character(defines):
    """将定义按角色分组整理。

    返回: [(group_key, display_name, [def, ...]), ...]
      group_key 为角色名或 "_system"
    """
    # 收集所有角色名
    characters = set()
    char_def_map = {}
    for d in defines:
        if d["type"] == "character":
            characters.add(d["name"])
            char_def_map[d["name"]] = d

    groups = {}        # {group_key: [def, ...]}
    virtual_groups = {}  # {(group_key, display_name): [def, ...]}
    ungrouped = []

    for d in defines:
        if d["type"] == "character":
            groups.setdefault(d["name"], []).append(d)
            continue

        # 分类前缀优先：首词命中 bg 等分类标记，直接归入对应虚拟组
        first_token = d["name"].lower().replace(" ", "_").split("_")[0]
        cat = _CATEGORY_FIRST_TOKENS.get(first_token)
        if cat:
            vkey, vlabel = cat
            virtual_groups.setdefault((vkey, vlabel), []).append(d)
            continue

        char = _match_def_to_character(d["name"], d["type"], characters)
        if char:
            groups.setdefault(char, []).append(d)
        else:
            virtual = _match_def_to_virtual(d["name"])
            if virtual:
                vkey, vlabel = virtual
                virtual_groups.setdefault((vkey, vlabel), []).append(d)
            else:
                ungrouped.append(d)

    result = []
    for char_name in sorted(groups.keys(), key=lambda n: n.lower()):
        # 构造显示名：角色中文名（英文名）
        char_display = char_name
        char_def = char_def_map.get(char_name)
        if char_def:
            m = re.search(r'Character\("([^"]*)"\)', char_def["line"])
            if m and m.group(1) != char_name:
                char_display = f"{m.group(1)}（{char_name}）"
        result.append((char_name, char_display, groups[char_name]))

    # 虚拟分组（如玩家）排在角色之后、系统之前
    for (vkey, vlabel) in sorted(virtual_groups.keys(), key=lambda k: k[1]):
        result.append((vkey, vlabel, virtual_groups[(vkey, vlabel)]))

    if ungrouped:
        result.append(("_system", "系统/全局定义", ungrouped))

    return result


def _format_defines_output(grouped_defines):
    """格式化分组的定义为 .rpy 文件内容。"""
    lines = ["# 由 excel_to_rpy.py 定义管理器自动生成", ""]

    for group_key, display_name, defs in grouped_defines:
        lines.append("")
        lines.append("=" * 40)
        lines.append(display_name)
        lines.append("=" * 40)
        lines.append("")

        # 排序：角色定义 → 变量（按名排序）→ 图片（按名排序）
        char_items = [d for d in defs if d["type"] == "character"]
        var_items = sorted(
            [d for d in defs if d["type"] == "variable"], key=lambda x: x["name"].lower()
        )
        img_items = sorted(
            [d for d in defs if d["type"] == "image"], key=lambda x: x["name"].lower()
        )

        for d in char_items + var_items + img_items:
            lines.append(d["line"])
        lines.append("")

    return "\n".join(lines) + "\n"


def append_defines_from_excel(excel_path):
    """从 Excel 提取定义，追加到 defines.rpy（不覆盖已有定义）。"""
    defines_rpy = _find_defines_rpy()
    if not defines_rpy:
        return  # 静默跳过

    new_defines = scan_excel_for_defines(excel_path)
    if not new_defines:
        return

    existing = parse_existing_defines(str(defines_rpy))

    to_add = [d for d in new_defines if d["name"] not in existing]
    if not to_add:
        print(f"[INFO] 定义管理：{len(new_defines)} 个定义全部已存在在 defines.rpy")
        return

    defines_rpy.parent.mkdir(parents=True, exist_ok=True)
    if not defines_rpy.exists():
        defines_rpy.write_text(
            "# 由 excel_to_rpy.py 定义管理器自动生成\n\n", encoding="utf-8"
        )

    with open(str(defines_rpy), "a", encoding="utf-8") as f:
        f.write("\n")
        for d in to_add:
            f.write(d["line"] + "\n")

    print(f"[INFO] 定义管理：追加了 {len(to_add)} 个新定义到 {defines_rpy.name}")
    skipped = len(new_defines) - len(to_add)
    if skipped:
        print(f"       {skipped} 个定义已存在，跳过。")


def _clean_excel_defines(excel_path, dry_run=False):
    """删除 Excel 中所有 define_character/define_variable/define_image 行。"""
    try:
        wb = load_workbook(excel_path)
    except Exception:
        return 0

    removed = 0
    for sheet in wb.worksheets:
        if sheet.sheet_state == "hidden" or sheet.title.startswith("_dd"):
            continue
        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2):
            if len(row) <= COL_CMD:
                continue
            raw_cmd = str(row[COL_CMD].value or "").strip()
            cmd = _parse_cmd(raw_cmd)
            if cmd in OLD_CMD_ALIASES:
                cmd = OLD_CMD_ALIASES[cmd]
            if cmd in ("define_character", "define_variable", "define_image"):
                rows_to_delete.append(row[0].row)

        for row_num in reversed(rows_to_delete):
            if not dry_run:
                sheet.delete_rows(row_num)
            removed += 1

    if removed and not dry_run:
        _safe_save_workbook(wb, excel_path, "Excel（清理后）")
    return removed


def _clean_rpy_defines(rpy_path, dry_run=False):
    """删除 .rpy 中的顶层 define/default/image 行。排除 defines.rpy 自身和 Ren'Py 系统文件。"""
    path_obj = Path(rpy_path)
    if path_obj.name.lower() == "defines.rpy":
        return 0
    if path_obj.name.lower() in _RENPY_SYSTEM_FILES:
        return 0

    try:
        content = path_obj.read_text(encoding="utf-8-sig")
    except Exception:
        return 0

    lines = content.splitlines()
    new_lines = []
    removed = 0

    for line in lines:
        stripped = line.strip()
        if (re.match(r'^\s*define\s+\w+\s*=\s*Character\b', stripped)
                or re.match(r'^\s*default\s+\w+\s*=', stripped)
                or re.match(r'^\s*image\s+\S', stripped)):
            removed += 1
            continue
        new_lines.append(line)

    if removed and not dry_run:
        result = "\n".join(new_lines) + "\n"
        _safe_write_text(rpy_path, result)
    return removed


def rebuild_all_defines(backup=False, cleanup=False, dry_run=False):
    """全量扫描项目中的 Excel 和 .rpy 文件，重建 defines.rpy。

    参数:
        backup: 是否备份旧 defines.rpy（默认 False）
        cleanup: 是否清理源文件中的定义行（默认 False）
        dry_run: 是否仅预览（默认 False）
    """
    defines_rpy = _find_defines_rpy()
    if not defines_rpy:
        game_dir = _find_game_dir()
        if game_dir:
            defines_rpy = game_dir / "defines.rpy"
        else:
            print("[ERROR] 未找到 game/ 目录，无法确定 defines.rpy 位置。")
            return

    root = _find_project_root()
    game_dir = _find_game_dir()

    print(f"\n[INFO] 项目根目录: {root}")
    print(f"[INFO] game 目录: {game_dir}")

    all_defines = []

    # 扫描 Excel 文件
    excel_files = []
    for ext in ("*.xlsx", "*.xlsm"):
        excel_files.extend(root.glob(ext))
        excel_files.extend(root.glob(f"**/{ext}"))
    excel_files = list(set(excel_files))

    print(f"\n[INFO] 扫描 {len(excel_files)} 个 Excel 文件...")
    for xlsx in excel_files:
        defs = scan_excel_for_defines(str(xlsx))
        if defs:
            print(f"  {xlsx.name}: {len(defs)} 个定义")
            all_defines.extend(defs)

    # 扫描 .rpy 文件（排除 defines.rpy）
    rpy_files = []
    if game_dir and game_dir.is_dir():
        rpy_files = list(game_dir.glob("**/*.rpy"))
        rpy_files = [f for f in rpy_files if f.name.lower() != "defines.rpy"]

    print(f"\n[INFO] 扫描 {len(rpy_files)} 个 .rpy 文件...")
    for rpy in rpy_files:
        defs = scan_rpy_for_defines(str(rpy))
        if defs:
            print(f"  {rpy.name}: {len(defs)} 个定义")
            all_defines.extend(defs)

    if not all_defines:
        print("[INFO] 未找到任何定义。")
        return

    # 去重：Excel 优先于 .rpy；同一来源内取第一个
    seen = {}
    conflicts = []
    for d in all_defines:
        if d["name"] in seen:
            existing = seen[d["name"]]
            if d["source"].startswith("excel:") and not existing["source"].startswith("excel:"):
                # Excel 覆盖 .rpy
                conflicts.append((d["name"], existing["source"], d["source"]))
                seen[d["name"]] = d
            elif existing["source"].startswith("excel:") and not d["source"].startswith("excel:"):
                # .rpy 被 Excel 覆盖，记录冲突
                conflicts.append((d["name"], d["source"], existing["source"]))
            else:
                conflicts.append((d["name"], existing["source"], d["source"]))
        else:
            seen[d["name"]] = d

    if conflicts:
        print(f"\n[INFO] 定义冲突（以 Excel 版本为准，{len(conflicts)} 个）:")
        for name, loser, winner in conflicts[:30]:
            print(f"  {name}: {loser} → {winner}")
        if len(conflicts) > 30:
            print(f"  ... 还有 {len(conflicts) - 30} 个冲突")

    defines_list = list(seen.values())
    grouped = _group_defines_by_character(defines_list)
    output = _format_defines_output(grouped)

    if dry_run:
        print(f"\n{'=' * 40}")
        print(f"  DRY-RUN 预览 — 将写入 {len(defines_list)} 个定义（{len(grouped)} 组）")
        print(f"{'=' * 40}")
        print(output)
        return

    # 备份
    if backup and defines_rpy.exists():
        import shutil as _shutil
        backup_path = defines_rpy.with_suffix(".rpy.bak")
        _shutil.copy2(str(defines_rpy), str(backup_path))
        print(f"[INFO] 已备份到: {backup_path}")

    # 写入
    defines_rpy.parent.mkdir(parents=True, exist_ok=True)
    defines_rpy.write_text(output, encoding="utf-8")
    print(f"\n[OK] defines.rpy 已重建: {len(defines_list)} 个定义, {len(grouped)} 组")

    # 清理源文件
    if cleanup:
        if dry_run:
            print("\n[DRY-RUN] 将清理源文件中的定义行...")
        else:
            print("\n[INFO] 清理源文件中的定义行...")

        excel_removed = 0
        for xlsx in excel_files:
            n = _clean_excel_defines(str(xlsx), dry_run)
            if n:
                excel_removed += n
                tag = "[DRY-RUN]" if dry_run else ""
                print(f"  {tag} {xlsx.name}: 删除 {n} 行")

        rpy_removed = 0
        for rpy_file in rpy_files:
            n = _clean_rpy_defines(str(rpy_file), dry_run)
            if n:
                rpy_removed += n
                tag = "[DRY-RUN]" if dry_run else ""
                print(f"  {tag} {rpy_file.name}: 删除 {n} 行")

        print(f"\n  合计: Excel {excel_removed} 行 + .rpy {rpy_removed} 行")


# ── 定义管理交互菜单 ──────────────────────────────────────

def _interactive_define_append():
    """交互式追加定义"""
    print("\n" + "-" * 40)
    print("  追加新定义")
    print("-" * 40)
    path = input("请输入 Excel 文件路径（可直接拖拽文件到此处）：\n> ").strip()
    path = _strip_quotes(path)
    if not path:
        print("[ERROR] 未输入路径。")
        return
    if not os.path.isfile(path):
        print(f"[ERROR] 文件不存在：{path}")
        return
    if not path.lower().endswith((".xlsx", ".xlsm")):
        print(f"[ERROR] 不是 .xlsx 文件。")
        return

    defines_rpy = _find_defines_rpy()
    if not defines_rpy:
        game_dir = _find_game_dir()
        if game_dir:
            defines_rpy = game_dir / "defines.rpy"
        else:
            print("[ERROR] 未找到 game/ 目录，无法确定 defines.rpy 位置。")
            return

    new_defines = scan_excel_for_defines(path)
    if not new_defines:
        print("[INFO] 未找到任何定义。")
        return

    existing = parse_existing_defines(str(defines_rpy))
    to_add = [d for d in new_defines if d["name"] not in existing]

    print(f"\n发现 {len(new_defines)} 个定义，其中 {len(to_add)} 个是新定义：")
    for d in to_add:
        print(f"  + {d['line']}")
    skipped = len(new_defines) - len(to_add)
    if skipped:
        print(f"\n{skipped} 个定义已存在，将跳过。")

    if not to_add:
        return

    confirm = input(f"\n确认追加 {len(to_add)} 个定义到 defines.rpy？(Y/n)：").strip().lower()
    if confirm and confirm != "y":
        print("已取消。")
        return

    defines_rpy.parent.mkdir(parents=True, exist_ok=True)
    if not defines_rpy.exists():
        defines_rpy.write_text(
            "# 由 excel_to_rpy.py 定义管理器自动生成\n\n", encoding="utf-8"
        )

    with open(str(defines_rpy), "a", encoding="utf-8") as f:
        f.write("\n")
        for d in to_add:
            f.write(d["line"] + "\n")

    print(f"[OK] 已追加 {len(to_add)} 个定义到 {defines_rpy.name}")


def _interactive_define_rebuild():
    """交互式重建定义"""
    print("\n" + "-" * 40)
    print("  重建全部定义")
    print("-" * 40)

    backup = input("是否备份旧 defines.rpy？(y/N)：").strip().lower() == "y"
    cleanup = input("是否清理源文件中的定义行？(y/N)：").strip().lower() == "y"
    dry_run = input("是否预览变更（dry-run）？(y/N)：").strip().lower() == "y"

    if not dry_run:
        msg = "确认执行重建"
        if backup:
            msg += "（将备份）"
        if cleanup:
            msg += "（将清理源文件）"
        confirm = input(f"\n{msg}？(Y/n)：").strip().lower()
        if confirm and confirm != "y":
            print("已取消。")
            return

    rebuild_all_defines(backup=backup, cleanup=cleanup, dry_run=dry_run)


def _interactive_define_manager():
    """定义管理子菜单"""
    while True:
        print()
        print("-" * 40)
        print("  定义管理（Define Manager）")
        print("-" * 40)
        print("  1. 追加新定义（从指定 Excel 提取新定义，追加到 defines.rpy）")
        print("  2. 重建全部定义（全量扫描所有源文件，重新生成 defines.rpy）")
        print("  3. 返回主菜单")
        print("-" * 40)
        choice = input("请选择 (1/2/3)：").strip()
        if choice == "1":
            _interactive_define_append()
            _pause()
        elif choice == "2":
            _interactive_define_rebuild()
            _pause()
        elif choice == "3":
            break
        else:
            print("[ERROR] 无效选择。")


# ── 入口 ───────────────────────────────────────────────────

def _strip_quotes(s: str) -> str:
    s = s.strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
        s = s[1:-1]
    return s


def _pause():
    input("\n按 Enter 键继续...")


def _interactive_convert():
    print("\n" + "-" * 40)
    path = input("请输入 Excel 文件路径（可直接拖拽文件到此处）：\n> ").strip()
    path = _strip_quotes(path)
    if not path:
        print("[ERROR] 未输入路径，返回主菜单。")
        return
    if not os.path.isfile(path):
        print(f"[ERROR] 文件不存在：{path}")
        return
    if not path.lower().endswith(".xlsx"):
        print(f"[ERROR] 不是 .xlsx 文件，请拖拽 Excel 表格。")
        print(f"       .rpy 文件请用 rpy_to_excel.py 打开。")
        return
    default_out = str(Path(path).with_suffix(".rpy"))
    out = input(f"输出路径（直接回车使用默认：{default_out}）：\n> ").strip()
    out = _strip_quotes(out) if out else default_out
    print("\n校验表格...")
    issues = check_excel(path)
    _print_check_result(issues)
    if any(i[1] == "error" for i in issues):
        go = input("\n存在错误，是否继续转换？(y/N)：").strip().lower()
        if go != "y":
            return
    convert_excel_to_rpy(path, out)


def _interactive_template():
    print("\n" + "-" * 40)
    default_path = "renpy_script_template.xlsx"
    path = input(f"模板保存路径（直接回车使用默认：{default_path}）：\n> ").strip()
    path = _strip_quotes(path) if path else default_path
    generate_template(path)


def _interactive_check():
    print("\n" + "-" * 40)
    path = input("请输入 Excel 文件路径（可直接拖拽文件到此处）：\n> ").strip()
    path = _strip_quotes(path)
    if not path:
        print("[ERROR] 未输入路径，返回主菜单。")
        return
    if not os.path.isfile(path):
        print(f"[ERROR] 文件不存在：{path}")
        return
    if not path.lower().endswith(".xlsx"):
        print(f"[ERROR] 不是 .xlsx 文件。")
        return
    issues = check_excel(path)
    _print_check_result(issues, max_show=50)


def _interactive_blank():
    """交互式生成空白模板"""
    print("\n" + "-" * 40)
    default_path = "renpy_script_blank.xlsx"
    path = input(f"空白模板保存路径（直接回车使用默认：{default_path}）：\n> ").strip()
    path = _strip_quotes(path) if path else default_path
    generate_blank_template(path)


def interactive_mode():
    while True:
        print()
        print("=" * 40)
        print("  Ren'Py Excel → .rpy 转换工具")
        print("=" * 40)
        print("  1. 转换 Excel 为 .rpy")
        print("  2. 校验 Excel 表格（不转换）")
        print("  3. 生成 Excel 模板（含示例和教学）")
        print("  4. 生成空白 Excel 模板（仅表头）")
        print("  5. 定义管理（Define Manager）")
        print("  6. 退出")
        print("-" * 40)
        choice = input("请选择 (1/2/3/4/5/6)：").strip()
        if choice == "1":
            _interactive_convert()
            _pause()
        elif choice == "2":
            _interactive_check()
            _pause()
        elif choice == "3":
            _interactive_template()
            _pause()
        elif choice == "4":
            _interactive_blank()
            _pause()
        elif choice == "5":
            _interactive_define_manager()
        elif choice == "6":
            print("再见！")
            break
        else:
            print("[ERROR] 无效选择，请输入 1、2、3、4、5 或 6。")


def _fix_working_directory():
    """如果当前工作目录是系统目录（如 C:\\Windows\\System32），切换到脚本所在目录。"""
    cwd = os.getcwd()
    windir = os.environ.get("SystemRoot", r"C:\Windows")
    windir = os.path.normpath(windir).lower()
    if os.path.normpath(cwd).lower().startswith(windir + os.sep) or os.path.normpath(cwd).lower() == windir:
        script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        os.chdir(script_dir)
        print(f"⚠️  已从系统目录（{cwd}）切换到脚本目录：{script_dir}\n")


def main():
    _fix_working_directory()
    if len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description="Ren'Py Excel → .rpy 转换工具")
        parser.add_argument("input", nargs="?", help="输入的 .xlsx 文件路径")
        parser.add_argument("-o", "--output", help="输出的 .rpy 文件路径")
        parser.add_argument("--template", action="store_true", help="生成 Excel 模板")
        parser.add_argument("--blank", action="store_true", help="生成空白 Excel 模板（无示例数据）")
        parser.add_argument("--template-path", default="renpy_script_template.xlsx", help="模板保存路径")
        parser.add_argument("--check", action="store_true", help="仅校验表格，不转换")
        parser.add_argument("--rebuild-defines", action="store_true", help="全量扫描源文件，重建 defines.rpy")
        parser.add_argument("--backup", action="store_true", help="重建 defines.rpy 前备份（配合 --rebuild-defines）")
        parser.add_argument("--no-cleanup", action="store_true", help="重建时不清理源文件中的定义行（配合 --rebuild-defines）")
        parser.add_argument("--dry-run", action="store_true", help="仅预览重建结果，不写入（配合 --rebuild-defines）")
        args = parser.parse_args()

        if args.rebuild_defines:
            rebuild_all_defines(
                backup=args.backup,
                cleanup=not args.no_cleanup,
                dry_run=args.dry_run,
            )
            return

        if args.template:
            generate_template(args.template_path)
            return

        if args.blank:
            generate_blank_template(args.template_path)
            return

        if not args.input:
            parser.print_help()
            print("\n提示：使用 --template 生成模板，或拖入 .xlsx 文件转换。")
            input("\n按 Enter 键退出...")
            return

        input_path = _strip_quotes(args.input)
        if not os.path.isfile(input_path):
            print(f"\u274c 文件不存在：{input_path}")
            input("\n按 Enter 键退出...")
            sys.exit(1)

        output_path = args.output if args.output else str(Path(input_path).with_suffix(".rpy"))

        print("校验表格...")
        issues = check_excel(input_path)
        _print_check_result(issues)

        if args.check:
            return
        if any(i[1] == "error" for i in issues):
            print("\n存在错误，使用 --check 仅校验不转换，或添加 -o 强制输出。")
            print("继续转换中...")
        convert_excel_to_rpy(input_path, output_path)
        try:
            input("\n按 Enter 键退出...")
        except (EOFError, OSError):
            pass
        return

    # 交互模式
    try:
        interactive_mode()
    except KeyboardInterrupt:
        print("\n\n再见！")
    except Exception as e:
        print(f"\n\u274c 发生错误：{e}")
        input("\n按 Enter 键退出...")


if __name__ == "__main__":
    main()
